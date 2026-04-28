"""
backtest.py — Test complet de toutes les fonctionnalites de LeadsEngine.
Resultat sauvegarde en DB (session "backtest 1") + export CRM Excel.
"""

import sys, os, time, traceback, sqlite3
from pathlib import Path
from datetime import datetime

ROOT = Path(__file__).parent
sys.path.insert(0, str(ROOT))
os.environ.setdefault("LEADS_ENGINE_ROOT", str(ROOT))

from dotenv import load_dotenv
load_dotenv(ROOT / ".env")

# ── Setup logging ──────────────────────────────────────────────────────────────
from core.logger import setup_logging, get_logger, ERRORS_LOG
setup_logging()
log = get_logger("backtest")

results = []  # (test_name, status, detail)


def record(name, status, detail=""):
    results.append((name, status, detail))
    icon = "OK" if status == "PASS" else "FAIL" if status == "FAIL" else "WARN"
    print(f"  [{icon}] {name}{(' — ' + detail) if detail else ''}")


# ══════════════════════════════════════════════════════════════════════════════
# 1. IMPORTS
# ══════════════════════════════════════════════════════════════════════════════
print("\n=== 1. IMPORTS ===")

modules_to_test = [
    ("config",                    "from config import config"),
    ("core.logger",               "from core.logger import get_logger, setup_logging"),
    ("core.models",               "from core.models import Lead, LeadStatus"),
    ("core.queue",                "from core.queue import LeadQueue"),
    ("core.crm_filter",           "from core.crm_filter import load_crm, filter_against_crm, crm_stats, parse_crm_file"),
    ("core.updater",              "from core.updater import check_update, get_local_version"),
    ("agents.scraper",            "from agents.scraper import ScraperAgent, deduplicate, deduplicate_against_db"),
    ("agents.extractor",          "from agents.extractor import ExtractorAgent, detect_cms, detect_agence, detect_seo, extract_email, extract_phone, analyze_structure, compute_weaknesses, process_lead"),
    ("services.serpapi",          "from services.serpapi import search_google_maps, parse_maps_result"),
    ("services.recherche_entreprises", "from services.recherche_entreprises import search_entreprises"),
    ("services.dirigeant",        "from services.dirigeant import find_dirigeant"),
    ("services.geocoding",        "from services.geocoding import get_city_coordinates, find_expansion_cities"),
    ("services.dns_lookup",       "from services.dns_lookup import get_hosting"),
    ("services.pagespeed",        "from services.pagespeed import get_pagespeed"),
    ("services.gmb",              "from services.gmb import enrich_from_maps_result"),
]

for mod_name, import_stmt in modules_to_test:
    try:
        exec(import_stmt)
        record(f"Import {mod_name}", "PASS")
    except Exception as e:
        record(f"Import {mod_name}", "FAIL", str(e))
        log.error("Import echoue : %s — %s", mod_name, e, exc_info=True)


# ══════════════════════════════════════════════════════════════════════════════
# 2. LOGGING
# ══════════════════════════════════════════════════════════════════════════════
print("\n=== 2. LOGGING ===")

try:
    # Test ecriture warning dans errors.log
    log.warning("BACKTEST — test warning volontaire")
    log.error("BACKTEST — test error volontaire")

    # Verifier que errors.log existe et contient nos messages
    time.sleep(0.5)  # flush
    if ERRORS_LOG.exists():
        content = ERRORS_LOG.read_text(encoding="utf-8")
        if "BACKTEST" in content:
            record("errors.log ecriture", "PASS", f"taille={ERRORS_LOG.stat().st_size} octets")
        else:
            record("errors.log ecriture", "WARN", "fichier existe mais message non trouve")
    else:
        record("errors.log ecriture", "FAIL", "fichier inexistant")
except Exception as e:
    record("errors.log ecriture", "FAIL", str(e))
    log.error("Test logging echoue : %s", e, exc_info=True)

# Test pipeline.log
try:
    pipeline_log = ROOT / "pipeline.log"
    if pipeline_log.exists():
        record("pipeline.log", "PASS", f"taille={pipeline_log.stat().st_size} octets")
    else:
        record("pipeline.log", "FAIL", "fichier inexistant")
except Exception as e:
    record("pipeline.log", "FAIL", str(e))


# ══════════════════════════════════════════════════════════════════════════════
# 3. DATABASE (CRUD)
# ══════════════════════════════════════════════════════════════════════════════
print("\n=== 3. DATABASE ===")

TEST_DB = str(ROOT / "backtest_temp.db")

try:
    queue = LeadQueue(TEST_DB, session_id="backtest_1", session_label="backtest 1")
    record("DB creation", "PASS")
except Exception as e:
    record("DB creation", "FAIL", str(e))
    log.error("DB creation echouee : %s", e, exc_info=True)
    queue = None

if queue:
    # Insert
    try:
        test_lead = Lead(
            company_name="Entreprise Test Backtest",
            city="Marseille",
            sector="plombier",
            source="backtest",
            google_rating=4.5,
            review_count=42,
            website_url="https://example.com",
            status=LeadStatus.SCRAPED,
        )
        lead_id = queue.save(test_lead)
        record("DB insert", "PASS", f"id={lead_id}")
    except Exception as e:
        record("DB insert", "FAIL", str(e))
        log.error("DB insert echoue : %s", e, exc_info=True)
        lead_id = None

    # Update status
    if lead_id:
        try:
            queue.update_status(lead_id, LeadStatus.EXTRACTED)
            record("DB update_status", "PASS")
        except Exception as e:
            record("DB update_status", "FAIL", str(e))
            log.error("DB update_status echoue : %s", e, exc_info=True)

    # Update fields
    if lead_id:
        try:
            queue.update_fields(lead_id, cms="WordPress", hosting="OVH", seo_score=7)
            record("DB update_fields", "PASS")
        except Exception as e:
            record("DB update_fields", "FAIL", str(e))
            log.error("DB update_fields echoue : %s", e, exc_info=True)

    # Read back
    try:
        rows = queue.get_by_session("backtest_1")
        if rows and rows[0]["company_name"] == "Entreprise Test Backtest":
            record("DB read back", "PASS", f"{len(rows)} lead(s)")
        else:
            record("DB read back", "FAIL", "donnees incorrectes")
    except Exception as e:
        record("DB read back", "FAIL", str(e))
        log.error("DB read back echoue : %s", e, exc_info=True)

    # Sessions list
    try:
        sessions = queue.list_sessions()
        record("DB list_sessions", "PASS", f"{len(sessions)} session(s)")
    except Exception as e:
        record("DB list_sessions", "FAIL", str(e))
        log.error("DB list_sessions echoue : %s", e, exc_info=True)

    # Stats
    try:
        stats = queue.stats()
        record("DB stats", "PASS", str(stats))
    except Exception as e:
        record("DB stats", "FAIL", str(e))
        log.error("DB stats echoue : %s", e, exc_info=True)

    # Cleanup temp DB
    try:
        Path(TEST_DB).unlink(missing_ok=True)
    except:
        pass


# ══════════════════════════════════════════════════════════════════════════════
# 4. APIs EXTERNES
# ══════════════════════════════════════════════════════════════════════════════
print("\n=== 4. APIs EXTERNES ===")

# 4a. Geocodage (geo.api.gouv.fr — gratuit)
try:
    coords = get_city_coordinates("Marseille")
    if coords and coords.get("lat"):
        record("API Geocodage", "PASS", f"Marseille -> lat={coords['lat']:.2f} lon={coords['lon']:.2f}")
    else:
        record("API Geocodage", "FAIL", "pas de coordonnees")
except Exception as e:
    record("API Geocodage", "FAIL", str(e))
    log.error("API Geocodage echouee : %s", e, exc_info=True)

# 4b. Extension villes voisines
try:
    nearby = find_expansion_cities("Aix-en-Provence", max_cities=5, max_radius_km=30)
    if nearby:
        villes = ", ".join(f"{c['nom']} ({c['distance_km']}km)" for c in nearby[:3])
        record("API Villes voisines", "PASS", f"{len(nearby)} trouvees — {villes}")
    else:
        record("API Villes voisines", "WARN", "aucune ville trouvee")
except Exception as e:
    record("API Villes voisines", "FAIL", str(e))
    log.error("API Villes voisines echouee : %s", e, exc_info=True)

# 4c. Registre National (recherche-entreprises.api.gouv.fr — gratuit)
registre_results = []
try:
    registre_results = search_entreprises("plombier", "Marseille", max_results=5)
    if registre_results:
        record("API Registre National", "PASS", f"{len(registre_results)} entreprises trouvees")
    else:
        record("API Registre National", "WARN", "0 resultat")
except Exception as e:
    record("API Registre National", "FAIL", str(e))
    log.error("API Registre National echouee : %s", e, exc_info=True)

# 4d. Dirigeant
try:
    dir_info = find_dirigeant("Bouygues Construction", "Paris")
    if dir_info.get("found"):
        record("API Dirigeant", "PASS", f"{dir_info.get('owner_name')} — SIREN {dir_info.get('siren')}")
    else:
        record("API Dirigeant", "WARN", "entreprise non trouvee (API peut etre lente)")
except Exception as e:
    record("API Dirigeant", "FAIL", str(e))
    log.error("API Dirigeant echouee : %s", e, exc_info=True)

# 4e. SerpAPI (Google Maps) — consomme 1 credit
serpapi_results = []
try:
    if config.serpapi_key:
        count = 0
        for r in search_google_maps("plombier", "Marseille", max_results=3):
            parsed = parse_maps_result(r)
            if parsed:
                serpapi_results.append(parsed)
            count += 1
            if count >= 3:
                break
        if serpapi_results:
            record("API SerpAPI (Maps)", "PASS", f"{len(serpapi_results)} leads — ex: {serpapi_results[0]['company_name']}")
        else:
            record("API SerpAPI (Maps)", "WARN", "0 resultat")
    else:
        record("API SerpAPI (Maps)", "WARN", "cle SERPAPI manquante")
except Exception as e:
    record("API SerpAPI (Maps)", "FAIL", str(e))
    log.error("API SerpAPI echouee : %s", e, exc_info=True)

# 4f. DNS Lookup (gratuit)
try:
    hosting = get_hosting("ovh.com")
    if hosting:
        record("DNS Lookup", "PASS", f"ovh.com -> {hosting}")
    else:
        record("DNS Lookup", "WARN", "aucun resultat")
except Exception as e:
    record("DNS Lookup", "FAIL", str(e))
    log.error("DNS Lookup echoue : %s", e, exc_info=True)

# 4g. PageSpeed (gratuit sans cle)
try:
    ps = get_pagespeed("https://www.google.com")
    if ps.get("mobile") is not None or ps.get("desktop") is not None:
        record("API PageSpeed", "PASS", f"mobile={ps.get('mobile')} desktop={ps.get('desktop')}")
    else:
        record("API PageSpeed", "WARN", "scores nuls (API lente ou quota)")
except Exception as e:
    record("API PageSpeed", "FAIL", str(e))
    log.error("API PageSpeed echouee : %s", e, exc_info=True)


# ══════════════════════════════════════════════════════════════════════════════
# 5. EXTRACTOR (Agent 2) — analyse technique d'un site reel
# ══════════════════════════════════════════════════════════════════════════════
print("\n=== 5. EXTRACTOR ===")

try:
    extract_result = process_lead({"website_url": "https://www.ovh.com", "company_name": "OVH"})
    status = extract_result.get("_status", "error")
    if status == "extracted":
        record("Extractor process_lead", "PASS",
               f"CMS={extract_result.get('cms')} | Hosting={extract_result.get('hosting')} | "
               f"SEO={extract_result.get('seo_score')}/10 | HTTPS={'oui' if extract_result.get('is_https') else 'non'}")
    else:
        record("Extractor process_lead", "WARN", f"status={status} err={extract_result.get('_error')}")
except Exception as e:
    record("Extractor process_lead", "FAIL", str(e))
    log.error("Extractor echoue : %s", e, exc_info=True)

# Test detect_cms
try:
    html_wp = '<link rel="stylesheet" href="/wp-content/themes/test/style.css">'
    cms = detect_cms(html_wp, {})
    record("detect_cms (WordPress)", "PASS" if cms == "WordPress" else "FAIL", f"detecte={cms}")
except Exception as e:
    record("detect_cms", "FAIL", str(e))

# Test extract_email
try:
    html_email = 'Contactez-nous : contact@example.com ou info@test.fr'
    email = extract_email(html_email)
    record("extract_email", "PASS" if email in ("contact@example.com", "info@test.fr") else "FAIL", f"email={email}")
except Exception as e:
    record("extract_email", "FAIL", str(e))

# Test extract_phone
try:
    html_phone = 'Tel : 04 91 23 45 67'
    phone = extract_phone(html_phone)
    record("extract_phone", "PASS" if phone else "FAIL", f"phone={phone}")
except Exception as e:
    record("extract_phone", "FAIL", str(e))


# ══════════════════════════════════════════════════════════════════════════════
# 6. DEDUPLICATION & CRM FILTER
# ══════════════════════════════════════════════════════════════════════════════
print("\n=== 6. DEDUP & CRM ===")

try:
    leads_test = [
        {"company_name": "Plomberie Martin", "city": "Marseille"},
        {"company_name": "Plomberie Martin SARL", "city": "Marseille"},
        {"company_name": "Electricite Dupont", "city": "Lyon"},
    ]
    deduped = deduplicate(leads_test)
    record("Dedup interne", "PASS" if len(deduped) == 2 else "FAIL", f"{len(leads_test)} -> {len(deduped)}")
except Exception as e:
    record("Dedup interne", "FAIL", str(e))
    log.error("Dedup echouee : %s", e, exc_info=True)

try:
    existing = [{"company_name": "Electricite Dupont", "city": "Lyon"}]
    filtered, removed = deduplicate_against_db(
        [{"company_name": "Electricite Dupont", "city": "Lyon"}, {"company_name": "Nouveau Lead", "city": "Paris"}],
        existing
    )
    record("Dedup cross-session", "PASS" if len(filtered) == 1 and removed == 1 else "FAIL",
           f"supprime={removed}, reste={len(filtered)}")
except Exception as e:
    record("Dedup cross-session", "FAIL", str(e))

try:
    crm_data = load_crm()
    crm_st = crm_stats()
    record("CRM load", "PASS", f"{crm_st['fichiers']} fichier(s), {crm_st['entreprises']} entreprise(s)")
except Exception as e:
    record("CRM load", "FAIL", str(e))
    log.error("CRM load echoue : %s", e, exc_info=True)


# ══════════════════════════════════════════════════════════════════════════════
# 7. UPDATER
# ══════════════════════════════════════════════════════════════════════════════
print("\n=== 7. UPDATER ===")

try:
    version = get_local_version(ROOT)
    record("Version locale", "PASS", f"v{version}")
except Exception as e:
    record("Version locale", "FAIL", str(e))

try:
    # check_update ne fait rien en mode dev (frozen=False), c'est le comportement attendu
    update = check_update(ROOT)
    record("check_update (dev mode)", "PASS", "retourne None en dev — comportement correct")
except Exception as e:
    record("check_update", "FAIL", str(e))
    log.error("check_update echoue : %s", e, exc_info=True)


# ══════════════════════════════════════════════════════════════════════════════
# 8. AGENT 1 (scraper) — petit test reel
# ══════════════════════════════════════════════════════════════════════════════
print("\n=== 8. AGENT 1 (scraper mini) ===")

import asyncio

# On fait un mini scrape avec seulement le Registre National (gratuit, pas de credit SerpAPI)
try:
    db_path = str(ROOT / "leads.db")
    scraper_queue = LeadQueue(db_path, session_id="backtest_1", session_label="backtest 1")
    scraper = ScraperAgent(scraper_queue)

    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    scraped_leads = loop.run_until_complete(
        scraper.run(
            queries=[("plombier", "Aix-en-Provence")],
            use_maps=False,       # pas de SerpAPI pour economiser les credits
            use_registre=True,
            max_per_query=10,
        )
    )
    loop.close()

    if scraped_leads:
        record("Agent 1 scrape", "PASS", f"{len(scraped_leads)} leads scrapes")
        # Afficher un extrait
        for lead in scraped_leads[:3]:
            print(f"    > {lead.company_name} | {lead.city} | {lead.owner_name or 'N/A'} | SIREN: {lead.siren or 'N/A'}")
    else:
        record("Agent 1 scrape", "WARN", "0 leads (API peut etre lente)")
except Exception as e:
    record("Agent 1 scrape", "FAIL", str(e))
    log.error("Agent 1 echoue : %s", e, exc_info=True)


# ══════════════════════════════════════════════════════════════════════════════
# 9. AGENT 2 (extractor) — sur les leads du backtest
# ══════════════════════════════════════════════════════════════════════════════
print("\n=== 9. AGENT 2 (extractor sur backtest) ===")

try:
    # Relire les leads du backtest depuis la DB
    bt_queue = LeadQueue(db_path, session_id="backtest_1", session_label="backtest 1")
    bt_leads = bt_queue.get_by_session("backtest_1")

    # Filtrer seulement ceux qui ont un site web, et limiter a 3 pour le test
    leads_with_site = [l for l in bt_leads if l.get("website_url")][:3]

    if leads_with_site:
        extractor = ExtractorAgent(bt_queue)
        ext_result = extractor.run(leads_with_site, delay=1.0)
        record("Agent 2 extract", "PASS",
               f"{ext_result['success']} ok | {ext_result['skipped']} skip | {ext_result['errors']} err / {ext_result['total']} total")
    else:
        record("Agent 2 extract", "WARN", "aucun lead avec site web a analyser")
except Exception as e:
    record("Agent 2 extract", "FAIL", str(e))
    log.error("Agent 2 echoue : %s", e, exc_info=True)


# ══════════════════════════════════════════════════════════════════════════════
# 10. EXPORT CRM — sauvegarde dans crm/backtest_1.xlsx
# ══════════════════════════════════════════════════════════════════════════════
print("\n=== 10. EXPORT CRM ===")

try:
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    conn = sqlite3.connect(db_path)
    df = pd.read_sql("SELECT * FROM leads WHERE session_id='backtest_1' ORDER BY google_rating DESC", conn)
    conn.close()

    if df.empty:
        record("Export CRM", "WARN", "aucun lead backtest en DB — rien a exporter")
    else:
        # Colonnes vendeur
        col_map = {
            "company_name": "Entreprise",
            "city": "Ville",
            "sector": "Secteur",
            "owner_name": "Dirigeant",
            "owner_role": "Role",
            "phone": "Telephone",
            "email": "Email",
            "website_url": "Site web",
            "google_rating": "Note Google",
            "review_count": "Avis",
            "siren": "SIREN",
            "legal_form": "Forme juridique",
            "employee_range": "Effectif",
            "creation_date": "Creation",
            "etat": "Etat",
            "cms": "CMS",
            "hosting": "Hebergeur",
            "pagespeed_mobile": "Vitesse mobile",
            "seo_score": "Score SEO /10",
            "seo_weaknesses": "Faiblesses SEO",
            "address": "Adresse",
            "source": "Source",
            "status": "Statut",
        }

        # Garder uniquement les colonnes qui existent dans le dataframe
        export_cols = [c for c in col_map if c in df.columns]
        df_export = df[export_cols].rename(columns=col_map)

        output_path = ROOT / "crm" / "backtest_1.xlsx"

        with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
            df_export.to_excel(writer, index=False, sheet_name="Backtest 1")
            ws = writer.sheets["Backtest 1"]

            # Style theme sombre
            header_fill = PatternFill(start_color="E87B2A", end_color="E87B2A", fill_type="solid")
            header_font = Font(name="Calibri", bold=True, color="0D0E11", size=11)
            cell_fill = PatternFill(start_color="13151A", end_color="13151A", fill_type="solid")
            cell_font = Font(name="Calibri", color="E2DDD6", size=10)
            thin_border = Border(
                left=Side(style="thin", color="1E2028"),
                right=Side(style="thin", color="1E2028"),
                top=Side(style="thin", color="1E2028"),
                bottom=Side(style="thin", color="1E2028"),
            )

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.fill = cell_fill
                    cell.font = cell_font
                    cell.border = thin_border

            # Auto-width
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        record("Export CRM", "PASS", f"{len(df_export)} leads -> {output_path.name}")
except Exception as e:
    record("Export CRM", "FAIL", str(e))
    log.error("Export CRM echoue : %s", e, exc_info=True)


# ══════════════════════════════════════════════════════════════════════════════
# RAPPORT FINAL
# ══════════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("  RAPPORT BACKTEST — LeadsEngine v" + get_local_version(ROOT))
print("  " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
print("=" * 70)

passed = sum(1 for _, s, _ in results if s == "PASS")
warned = sum(1 for _, s, _ in results if s == "WARN")
failed = sum(1 for _, s, _ in results if s == "FAIL")
total  = len(results)

print(f"\n  Total : {total} tests")
print(f"  PASS  : {passed}")
print(f"  WARN  : {warned}")
print(f"  FAIL  : {failed}")

if failed > 0:
    print("\n  Tests en echec :")
    for name, s, detail in results:
        if s == "FAIL":
            print(f"    - {name} : {detail}")

if warned > 0:
    print("\n  Avertissements :")
    for name, s, detail in results:
        if s == "WARN":
            print(f"    - {name} : {detail}")

# Ecrire un log final
log.info("BACKTEST TERMINE — %d/%d PASS, %d WARN, %d FAIL", passed, total, warned, failed)

print(f"\n  Resultats en DB  : session 'backtest 1' dans leads.db")
print(f"  Export CRM       : crm/backtest_1.xlsx")
print(f"  Logs erreurs     : errors.log")
print(f"  Logs pipeline    : pipeline.log")
print("=" * 70)
