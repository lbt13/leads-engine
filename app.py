import sys, asyncio, sqlite3, io, threading, re, time, subprocess
from datetime import datetime
from pathlib import Path

import pandas as pd
import streamlit as st
import nest_asyncio
nest_asyncio.apply()

# ROOT  = dossier des données (leads.db, crm/, .env…)
#         → %APPDATA%\LeadsEngine en exe, dossier source en dev
# _CODE = dossier du code Python (pour sys.path)
#         → _MEIPASS en exe (géré par PyInstaller), dossier source en dev
import os as _os
ROOT  = Path(_os.environ.get("LEADS_ENGINE_ROOT", str(Path(__file__).parent)))
_CODE = Path(getattr(sys, "_MEIPASS", str(ROOT)))
if str(_CODE) not in sys.path:
    sys.path.insert(0, str(_CODE))
from dotenv import load_dotenv; load_dotenv(ROOT / ".env")
from core.logger import get_logger, setup_logging
setup_logging()
log = get_logger("app")

from config import config
from core.queue import LeadQueue
from core.crm_filter import crm_stats, CRM_DIR, compare_against_crm, parse_crm_file
from core.user_config import load as load_user_config, set_crm, save as save_user_config
from core.crm_export import CRM_MAPPINGS, get_crm_list, export_crm_csv
from core.scoring import compute_lead_score, score_label
from core.mailer import is_gmail_configured, send_email, check_replies
from core.caller import is_twilio_configured, make_call_twilio, is_telnyx_configured, make_call_telnyx
from core.license import is_activated, is_pro, is_standard, get_tier, activate as activate_license, get_license_key, TIER_LABELS
from core.crm_push import PUSH_CAPABLE_CRMS, is_connected, test_connection, push_leads, hubspot_auth_url, hubspot_exchange_code, salesforce_auth_url, salesforce_exchange_code
from core.updater import check_update, download_and_install, launch_update_and_quit, get_local_version
from agents.scraper import ScraperAgent
from agents.extractor import ExtractorAgent

# Crée les dossiers et la DB dès le démarrage — garantit que tout existe
# indépendamment de si un scraping a déjà été fait ou non.
(ROOT / "crm").mkdir(parents=True, exist_ok=True)
(ROOT / "analyses_a2").mkdir(parents=True, exist_ok=True)
LeadQueue(str(ROOT / config.db_path))  # initialise leads.db si absent

st.set_page_config(page_title="Leads Engine", page_icon="◈", layout="wide", initial_sidebar_state="collapsed")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Epilogue:wght@300;400;500;600;700;800;900&family=IBM+Plex+Mono:wght@300;400;500&display=swap');
*,*::before,*::after{box-sizing:border-box}
html,body,[class*="css"],.stApp{font-family:'Epilogue',sans-serif !important;background-color:#0D0E11 !important;color:#E2DDD6 !important}
[data-testid="stSidebar"],[data-testid="collapsedControl"]{display:none !important}
.main .block-container{padding:0 !important;max-width:100% !important}
h1{font-family:'Epilogue',sans-serif !important;font-size:26px !important;font-weight:800 !important;letter-spacing:-0.8px !important;color:#F0EBE3 !important;margin-bottom:4px !important;line-height:1.1 !important}
h2{font-family:'Epilogue',sans-serif !important;font-size:15px !important;font-weight:700 !important;color:#C8C2BB !important}
[data-testid="stMetric"]{background:#13151A !important;border:1px solid #1E2028 !important;border-radius:8px !important;padding:14px 18px !important}
[data-testid="stMetric"]:hover{border-color:#2E3240 !important}
[data-testid="stMetricLabel"]{font-size:10px !important;font-weight:600 !important;text-transform:uppercase !important;letter-spacing:1.2px !important;color:#4A4D58 !important}
[data-testid="stMetricValue"]{font-family:'IBM Plex Mono',monospace !important;font-size:24px !important;font-weight:500 !important;color:#F0EBE3 !important;line-height:1.2 !important}
.stTextInput>div>div>input,.stSelectbox>div>div>div,.stMultiSelect>div>div>div{background-color:#13151A !important;border:1px solid #1E2028 !important;border-radius:6px !important;color:#E2DDD6 !important;font-family:'Epilogue',sans-serif !important}
.stTextInput>div>div>input:focus{border-color:#E87B2A !important;box-shadow:0 0 0 2px rgba(232,123,42,.12) !important}
.stTextInput label,.stSelectbox label,.stSlider label,.stMultiSelect label{font-size:10px !important;font-weight:700 !important;text-transform:uppercase !important;letter-spacing:1.2px !important;color:#4A4D58 !important}
.stButton>button[kind="primary"]{background:#E87B2A !important;color:#0D0E11 !important;border:none !important;border-radius:6px !important;font-family:'Epilogue',sans-serif !important;font-weight:700 !important;font-size:13px !important}
.stButton>button[kind="primary"]:hover{opacity:.88 !important}
.stButton>button[kind="secondary"]{background:#13151A !important;color:#C8C2BB !important;border:1px solid #1E2028 !important;border-radius:6px !important;font-family:'Epilogue',sans-serif !important;font-weight:600 !important;font-size:13px !important}
.stButton>button[kind="secondary"]:hover{border-color:#E87B2A !important;color:#E87B2A !important}
[data-testid="stDownloadButton"] button{background:#13151A !important;color:#C8C2BB !important;border:1px solid #1E2028 !important;border-radius:6px !important;font-family:'Epilogue',sans-serif !important;font-weight:600 !important;font-size:13px !important}
[data-testid="stDownloadButton"] button:hover{border-color:#E87B2A !important;color:#E87B2A !important}
[data-testid="stDataFrame"]{border:1px solid #1E2028 !important;border-radius:8px !important;overflow:hidden !important}
[data-testid="stDataFrame"] th{background-color:#13151A !important;color:#4A4D58 !important;font-family:'IBM Plex Mono',monospace !important;font-size:10px !important;text-transform:uppercase !important;letter-spacing:1px !important}
[data-testid="stDataFrame"] td{color:#C8C2BB !important;font-family:'IBM Plex Mono',monospace !important;font-size:12px !important;background-color:#0D0E11 !important}
[data-testid="stDataFrame"] tr:hover td{background-color:#13151A !important}
.stSuccess{background:#0A1A10 !important;border:1px solid #1A4228 !important;border-left:3px solid #2ECC71 !important;border-radius:6px !important;color:#7DFAB8 !important}
.stError{background:#1A0A0A !important;border:1px solid #4A1A1A !important;border-left:3px solid #E84C4C !important;border-radius:6px !important}
.stInfo{background:#0D0F1A !important;border:1px solid #1A1E3A !important;border-left:3px solid #4A7CFF !important;border-radius:6px !important;color:#8FA8FF !important}
.stWarning{background:#1A1200 !important;border:1px solid #3A2E00 !important;border-left:3px solid #E8C32A !important;border-radius:6px !important;color:#F5D87A !important}
.stToggle label,.stCheckbox label{color:#C8C2BB !important;font-size:14px !important}
.stProgress>div>div{background:#E87B2A !important;border-radius:2px !important}
.stCaption{color:#4A4D58 !important;font-family:'IBM Plex Mono',monospace !important;font-size:11px !important}
hr{border-color:#1E2028 !important;margin:20px 0 !important}
.stExpander{border:1px solid #1E2028 !important;border-radius:8px !important;background:#13151A !important}
[data-testid="stTabs"] [role="tablist"]{border-bottom:1px solid #1E2028 !important;gap:0 !important;background:transparent !important;padding:0 40px !important}
[data-testid="stTabs"] button[role="tab"]{font-family:'Epilogue',sans-serif !important;font-size:13px !important;font-weight:600 !important;color:#4A4D58 !important;padding:14px 22px !important;border-radius:0 !important;border-bottom:2px solid transparent !important;background:transparent !important}
[data-testid="stTabs"] button[role="tab"][aria-selected="true"]{color:#F0EBE3 !important;border-bottom:2px solid #E87B2A !important}
[data-testid="stTabs"] [role="tabpanel"]{padding:28px 40px 0 !important}
.section-lbl{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:2px;color:#4A4D58;margin-bottom:14px;padding-bottom:8px;border-bottom:1px solid #1E2028}
.a2-banner{background:linear-gradient(135deg,#13151A 0%,#1A1008 100%);border:1px solid #2E2010;border-left:3px solid #E87B2A;border-radius:8px;padding:14px 18px;margin-bottom:18px}
.a2-banner .t{font-size:13px;font-weight:700;color:#E87B2A;margin-bottom:2px}
.a2-banner .d{font-size:12px;color:#6A6560}
.export-box{margin-top:20px;padding:18px 22px;background:#13151A;border:1px solid #1E2028;border-radius:8px}
.export-lbl{font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:1.5px;color:#4A4D58;margin-bottom:14px}
</style>
""", unsafe_allow_html=True)


# ── Écran d'activation de licence ────────────────────────────────────────────
if not is_activated():
    st.markdown("""
    <div style="max-width:600px;margin:80px auto;text-align:center">
        <div style="font-size:32px;font-weight:900;color:#F0EBE3;letter-spacing:-1px;margin-bottom:4px">
            ◈ leads<span style="color:#E87B2A">.</span>engine
        </div>
        <div style="font-size:12px;color:#4A4D58;letter-spacing:2px;text-transform:uppercase;margin-bottom:40px">
            Activation de la licence
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div style="max-width:520px;margin:0 auto">
        <p style="color:#C8C2BB;font-size:14px;text-align:center;margin-bottom:30px">
            Entre ta clé de licence pour activer l'application.<br>
            <span style="color:#4A4D58;font-size:12px">Format : LE-STD-XXXX-XXXX ou LE-PRO-XXXX-XXXX</span>
        </p>
    </div>
    """, unsafe_allow_html=True)

    _, _c_lic, _ = st.columns([1, 2, 1])
    with _c_lic:
        _lic_input = st.text_input("Clé de licence", key="license_input", placeholder="LE-XXX-XXXX-XXXX")
        if st.button("Activer", type="primary", use_container_width=True, key="btn_activate"):
            if _lic_input.strip():
                _ok, _msg = activate_license(_lic_input)
                if _ok:
                    st.success(_msg)
                    time.sleep(1)
                    st.rerun()
                else:
                    st.error(_msg)
            else:
                st.error("Entre une clé de licence.")

    st.markdown("""
    <div style="max-width:520px;margin:30px auto 0;text-align:center">
        <a href="https://leadsengine.netlify.app" target="_blank"
           style="color:#E87B2A;font-size:13px;text-decoration:none;font-weight:600">
            Acheter une licence sur leadsengine.netlify.app
        </a>
    </div>
    """, unsafe_allow_html=True)

    st.stop()


# ── Tier courant ─────────────────────────────────────────────────────────────
_TIER = get_tier()
_IS_PRO = _TIER == "pro"

# ── Écran de configuration (1er lancement) ───────────────────────────────────
_user_cfg = load_user_config()
if not _user_cfg.get("setup_done"):
    st.markdown("""
    <div style="max-width:600px;margin:80px auto;text-align:center">
        <div style="font-size:32px;font-weight:900;color:#F0EBE3;letter-spacing:-1px;margin-bottom:4px">
            ◈ leads<span style="color:#E87B2A">.</span>engine
        </div>
        <div style="font-size:12px;color:#4A4D58;letter-spacing:2px;text-transform:uppercase;margin-bottom:40px">
            Configuration initiale
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    <div style="max-width:520px;margin:0 auto">
        <p style="color:#C8C2BB;font-size:14px;text-align:center;margin-bottom:30px">
            Choisis ton CRM pour que les exports soient directement importables.<br>
            Tu pourras changer ce choix plus tard dans les paramètres.
        </p>
    </div>
    """, unsafe_allow_html=True)

    _crm_options = [{"key": "default", "label": "Export standard (Excel/CSV)", "description": "Pas de CRM — exports classiques"}]
    _crm_options += get_crm_list()

    _cols_setup = st.columns(3)
    for i, opt in enumerate(_crm_options):
        with _cols_setup[i % 3]:
            _selected = st.button(
                f"{opt['label']}",
                key=f"setup_crm_{opt['key']}",
                use_container_width=True,
                type="primary" if opt["key"] == "default" else "secondary",
                help=opt["description"],
            )
            if _selected:
                set_crm(opt["key"])
                st.rerun()

    st.markdown("<div style='height:30px'></div>", unsafe_allow_html=True)
    _, _c_later, _ = st.columns([1, 1, 1])
    with _c_later:
        if st.button("Configurer plus tard", use_container_width=True, type="secondary", key="setup_later"):
            save_user_config({"crm": "default", "setup_done": True})
            st.rerun()

    st.stop()


# ── Helpers ───────────────────────────────────────────────────────────────────

def load_df(session_id=None):
    db = ROOT / config.db_path
    if not db.exists():
        return pd.DataFrame()
    conn = sqlite3.connect(db)
    q = ("SELECT * FROM leads WHERE session_id=? ORDER BY google_rating DESC"
         if session_id else "SELECT * FROM leads ORDER BY scraped_at DESC")
    df = pd.read_sql(q, conn, params=(session_id,) if session_id else None)
    conn.close()
    return df


def _autofit(writer):
    for sh in writer.sheets.values():
        for col in sh.columns:
            sh.column_dimensions[col[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in col) + 3, 38)


def to_excel_a1(df_a1) -> bytes:
    """Export Agent 1 uniquement — un seul onglet."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df_a1.to_excel(w, index=False, sheet_name="Leads")
        _autofit(w)
    return buf.getvalue()


def to_excel_combined(df, cols_a1: dict, cols_a2: dict) -> bytes:
    """
    Export Agent 1 + Agent 2 fusionnés dans un seul onglet.
    Toutes les colonnes des deux agents côte à côte sur chaque ligne.
    """
    # Colonnes A1 + A2 sans doublons (A1 en premier)
    all_cols = {**cols_a1, **{k: v for k, v in cols_a2.items() if k not in cols_a1}}
    cols_present = [c for c in all_cols if c in df.columns]
    df_export = df[cols_present].rename(columns=all_cols)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df_export.to_excel(w, index=False, sheet_name="Leads complets")
        _autofit(w)
    return buf.getvalue()


def to_excel_qualifies(df_qual, cols_a1: dict, cols_a2: dict) -> bytes:
    """Export leads qualifiés avec toutes les colonnes A1 + A2."""
    all_cols = {**cols_a1, **{k: v for k, v in cols_a2.items() if k not in cols_a1}}
    cols_present = [c for c in all_cols if c in df_qual.columns]
    df_export = df_qual[cols_present].rename(columns=all_cols)
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df_export.to_excel(w, index=False, sheet_name="Leads qualifiés")
        _autofit(w)
    return buf.getvalue()


def to_excel_multi(df_a1, df_a2=None, df_qual=None):
    """Conservé pour compatibilité — utilisé dans l'onglet global."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df_a1.to_excel(w, index=False, sheet_name="Infos générales")
        if df_a2 is not None and not df_a2.empty:
            df_a2.to_excel(w, index=False, sheet_name="Technique site web")
        if df_qual is not None and not df_qual.empty:
            df_qual.to_excel(w, index=False, sheet_name="Leads qualifiés")
        _autofit(w)
    return buf.getvalue()


def to_excel_vendeur(df) -> bytes:
    """Export vue vendeur — thème clair professionnel, épuré."""
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter

    cols_present = [c for c in COLS_VENDEUR if c in df.columns]
    df_exp = df[cols_present].rename(columns=COLS_VENDEUR).copy()

    # Styles — thème sombre
    fill_header  = PatternFill("solid", fgColor="E87B2A")
    fill_row_a   = PatternFill("solid", fgColor="0D0E11")
    fill_row_b   = PatternFill("solid", fgColor="13151A")
    font_header  = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    font_company = Font(name="Calibri", bold=True, color="F0EBE3", size=10)
    font_normal  = Font(name="Calibri", color="C8C2BB", size=10)
    font_seo_ok  = Font(name="Calibri", bold=True, color="2ECC71", size=10)
    font_seo_mid = Font(name="Calibri", bold=True, color="F39C12", size=10)
    font_seo_bad = Font(name="Calibri", bold=True, color="E74C3C", size=10)
    font_weak    = Font(name="Calibri", color="E87B2A", size=10)
    font_link    = Font(name="Calibri", color="5DADE2", underline="single", size=10)
    border_cell  = Border(
        top=Side(style="thin", color="1E2028"),
        bottom=Side(style="thin", color="1E2028"),
        left=Side(style="thin", color="1E2028"),
        right=Side(style="thin", color="1E2028"),
    )

    col_names   = list(df_exp.columns)
    seo_col     = col_names.index("Score SEO /10")  + 1 if "Score SEO /10"  in col_names else None
    weak_col    = col_names.index("Faiblesses SEO") + 1 if "Faiblesses SEO" in col_names else None
    name_col    = col_names.index("Entreprise")     + 1 if "Entreprise"     in col_names else None
    site_col    = col_names.index("Site web")       + 1 if "Site web"       in col_names else None

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df_exp.to_excel(w, index=False, sheet_name="Leads")
        ws = w.sheets["Leads"]

        # En-tête
        ws.row_dimensions[1].height = 22
        for cell in ws[1]:
            cell.fill      = fill_header
            cell.font      = font_header
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = border_cell

        # Données
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = fill_row_a if row_idx % 2 == 0 else fill_row_b
            ws.row_dimensions[row_idx].height = 18
            for cell in row:
                cell.fill      = fill
                cell.font      = font_normal
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                cell.border    = border_cell

            # Entreprise en gras
            if name_col:
                ws.cell(row=row_idx, column=name_col).font = font_company

            # Site web en lien
            if site_col:
                c = ws.cell(row=row_idx, column=site_col)
                if c.value:
                    c.font = font_link

            # Score SEO coloré
            if seo_col:
                c = ws.cell(row=row_idx, column=seo_col)
                if isinstance(c.value, (int, float)):
                    c.font = font_seo_ok if c.value >= 7 else (font_seo_mid if c.value >= 4 else font_seo_bad)
                    c.alignment = Alignment(horizontal="center", vertical="center")

            # Faiblesses en rouge discret
            if weak_col:
                c = ws.cell(row=row_idx, column=weak_col)
                if c.value:
                    c.font      = font_weak
                    c.alignment = Alignment(vertical="center", wrap_text=True)

        # Largeurs colonnes
        FIXED_WIDTHS = {
            "Entreprise": 28, "Ville": 16, "Secteur": 18, "Dirigeant": 20,
            "Rôle": 16, "Téléphone": 14, "Email": 26, "Site web": 30,
            "Note Google": 12, "Avis": 8, "Forme juridique": 16, "Effectif": 12,
            "CMS": 12, "Hébergeur": 14, "Vitesse mobile": 14,
            "Score SEO /10": 13, "Faiblesses SEO": 42, "Création": 12,
            "État": 10, "SIREN": 12, "Adresse": 28,
        }
        for col in ws.columns:
            header = ws.cell(row=1, column=col[0].column).value
            width  = FIXED_WIDTHS.get(header, 16)
            ws.column_dimensions[get_column_letter(col[0].column)].width = width

        # Figer en-tête + tab name
        ws.freeze_panes   = "A2"
        ws.sheet_view.showGridLines = False

    return buf.getvalue()


def fmt_session(s):
    return s.get("session_label") or s.get("session_id") or "—"


def is_qualifie(row) -> bool:
    """
    Lead qualifié = les 10 critères suivants sont tous remplis :
    nom, adresse, téléphone, ville, secteur,
    dirigeant, CMS, hébergeur, vitesse du site, présence/absence SEO keywords
    """
    adresse = row.get("address") or row.get("siege_adresse")
    return all([
        row.get("company_name"),
        adresse,
        row.get("phone"),
        row.get("city"),
        row.get("sector"),
        row.get("owner_name"),
        row.get("cms"),
        row.get("hosting"),
        row.get("pagespeed_mobile") is not None,
        row.get("has_seo_keywords") is not None,
    ])


def has_agent2(df):
    return "cms" in df.columns and df["cms"].notna().any()


def db_stats():
    db = ROOT / config.db_path
    if not db.exists():
        return {}
    try:
        conn = sqlite3.connect(db)
        r = {
            "total":    conn.execute("SELECT COUNT(*) FROM leads").fetchone()[0],
            "sessions": conn.execute("SELECT COUNT(DISTINCT session_id) FROM leads").fetchone()[0],
            "analyses": conn.execute("SELECT COUNT(*) FROM leads WHERE status='extracted'").fetchone()[0],
            "qualifies": conn.execute("""
                SELECT COUNT(*) FROM leads
                WHERE company_name IS NOT NULL
                AND (address IS NOT NULL OR siege_adresse IS NOT NULL)
                AND phone IS NOT NULL AND city IS NOT NULL AND sector IS NOT NULL
                AND owner_name IS NOT NULL AND cms IS NOT NULL AND hosting IS NOT NULL
                AND pagespeed_mobile IS NOT NULL AND has_seo_keywords IS NOT NULL
            """).fetchone()[0],
        }
        conn.close()
        return r
    except Exception:
        log.warning("Calcul stats DB echoue", exc_info=True)
        return {}


ANALYSES_A2_DIR = ROOT / "analyses_a2"
ANALYSES_A2_DB  = ANALYSES_A2_DIR / "analyses.db"


def _enrich_websites(sid: str, db_path: str, delay: float, result_holder: list, error_holder: list):
    """
    Cherche les sites web manquants via Google Maps (SerpAPI).
    Regroupe les entreprises par (secteur, ville) et fait UNE requête par groupe
    (même logique que le scraping) — beaucoup plus économe en crédits.
    """
    import re as _re
    from difflib import SequenceMatcher
    from services.serpapi import search_google_maps

    def _norm(s):
        return _re.sub(r"\s+", " ", _re.sub(r"[^\w\s]", "", str(s or "").lower())).strip()

    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        leads = [dict(r) for r in conn.execute(
            "SELECT id, company_name, city, sector FROM leads WHERE session_id=? AND (website_url IS NULL OR website_url='')",
            (sid,)
        ).fetchall()]
        conn.close()

        total = len(leads)
        found = 0

        # Regrouper par (secteur, ville) — une seule requête Maps par groupe
        # On retire le code postal (ex: "13100 Aix-en-Provence" → "aix en provence")
        def _city_key(city: str) -> str:
            return _norm(_re.sub(r"^\d{4,5}\s*", "", str(city or "")).strip())

        groups: dict[tuple, list] = {}
        for lead in leads:
            key = (_norm(lead.get("sector") or ""), _city_key(lead.get("city") or ""))
            groups.setdefault(key, []).append(lead)

        for (sector, city), group in groups.items():
            # Ignorer les groupes sans ville ou trop petits (< 3 entreprises)
            # — peu de chance de trouver les bonnes entreprises sans localisation précise
            if not city or len(group) < 3:
                continue

            # Construire le lookup nom → id pour ce groupe
            lookup = {_norm(l["company_name"]): l for l in group}

            query = sector or " ".join(l["company_name"] for l in group[:3])
            if not query:
                continue

            try:
                for raw in search_google_maps(query, city, max_results=20):
                    result_name = raw.get("title", "")
                    w = (raw.get("website") or raw.get("links", {}).get("website") or "").strip().rstrip("/")
                    if not w:
                        continue
                    # Chercher la meilleure correspondance dans ce groupe
                    best_key, best_ratio = None, 0.0
                    for k in lookup:
                        r = SequenceMatcher(None, _norm(result_name), k).ratio()
                        if r > best_ratio:
                            best_ratio, best_key = r, k
                    if best_ratio >= 0.55 and best_key:
                        lead_match = lookup.pop(best_key)
                        c2 = sqlite3.connect(db_path)
                        c2.execute("UPDATE leads SET website_url=?, status='scraped' WHERE id=?",
                                   (w, lead_match["id"]))
                        c2.commit()
                        c2.close()
                        found += 1
                    if not lookup:
                        break  # tous les leads du groupe sont appariés
            except Exception:
                log.warning("Enrichissement site web echoue pour un groupe", exc_info=True)

            time.sleep(delay)

        result_holder.append((found, total))
    except Exception as e:
        log.error("_enrich_websites erreur globale", exc_info=True)
        error_holder.append(e)


def _find_db_duplicates() -> tuple[int, int]:
    """
    Analyse leads.db et supprime les doublons (SIREN exact ou nom+ville ≥ 85%).
    Conserve l'entrée la plus riche pour chaque groupe.
    Retourne (nb_groupes, nb_supprimés).
    """
    import re as _re
    from difflib import SequenceMatcher

    def _norm(text):
        if not text: return ""
        text = str(text).lower().strip()
        for s in [" sarl", " sas", " eurl", " sa ", " sci ", " sasu"]:
            text = text.replace(s, " ")
        text = _re.sub(r"[^a-z0-9\s]", " ", text)
        return _re.sub(r"\s+", " ", text).strip()

    def _score(lead):
        s = 0
        if lead.get("session_id") and not lead["session_id"].startswith("crm_"):
            s += 1000
        for col in ["siren", "website_url", "owner_name", "phone", "email",
                    "address", "cms", "hosting", "seo_score", "pagespeed_mobile"]:
            if lead.get(col): s += 1
        return s

    conn = sqlite3.connect(str(ROOT / config.db_path))
    conn.row_factory = sqlite3.Row
    rows = [dict(r) for r in conn.execute("SELECT * FROM leads ORDER BY id").fetchall()]

    processed, groups = set(), []
    for i, a in enumerate(rows):
        if a["id"] in processed: continue
        na = _norm(a["company_name"])
        ca = _norm(a["city"] or "")
        sa = _re.sub(r"\D", "", str(a.get("siren") or ""))
        group = [a]
        for j, b in enumerate(rows):
            if i == j or b["id"] in processed or b["id"] in {x["id"] for x in group}:
                continue
            nb = _norm(b["company_name"])
            cb = _norm(b["city"] or "")
            sb = _re.sub(r"\D", "", str(b.get("siren") or ""))
            if sa and len(sa) == 9 and sa == sb:
                group.append(b); continue
            if ca and cb and ca != cb: continue
            if SequenceMatcher(None, na, nb).ratio() >= 0.85:
                group.append(b)
        if len(group) > 1:
            for x in group:
                processed.add(x["id"])
            groups.append(group)

    ids_to_delete = []
    for group in groups:
        sorted_group = sorted(group, key=_score, reverse=True)
        ids_to_delete.extend(x["id"] for x in sorted_group[1:])

    if ids_to_delete:
        conn.executemany("DELETE FROM leads WHERE id=?", [(i,) for i in ids_to_delete])
        conn.commit()
    conn.close()
    return len(groups), len(ids_to_delete)


def _count_db_duplicates() -> tuple[int, int]:
    """Compte les doublons sans rien supprimer. Retourne (nb_groupes, nb_en_trop)."""
    import re as _re
    from difflib import SequenceMatcher

    def _norm(text):
        if not text: return ""
        text = str(text).lower().strip()
        for s in [" sarl", " sas", " eurl", " sa ", " sci ", " sasu"]:
            text = text.replace(s, " ")
        text = _re.sub(r"[^a-z0-9\s]", " ", text)
        return _re.sub(r"\s+", " ", text).strip()

    conn = sqlite3.connect(str(ROOT / config.db_path))
    conn.row_factory = sqlite3.Row
    rows = [dict(r) for r in conn.execute("SELECT id, company_name, city, siren, session_id FROM leads ORDER BY id").fetchall()]
    conn.close()

    processed, nb_groupes, nb_en_trop = set(), 0, 0
    for i, a in enumerate(rows):
        if a["id"] in processed: continue
        na = _norm(a["company_name"])
        ca = _norm(a["city"] or "")
        sa = _re.sub(r"\D", "", str(a.get("siren") or ""))
        group = [a]
        for j, b in enumerate(rows):
            if i == j or b["id"] in processed or b["id"] in {x["id"] for x in group}:
                continue
            nb = _norm(b["company_name"])
            cb = _norm(b["city"] or "")
            sb = _re.sub(r"\D", "", str(b.get("siren") or ""))
            if sa and len(sa) == 9 and sa == sb:
                group.append(b); continue
            if ca and cb and ca != cb: continue
            if SequenceMatcher(None, na, nb).ratio() >= 0.85:
                group.append(b)
        if len(group) > 1:
            for x in group: processed.add(x["id"])
            nb_groupes += 1
            nb_en_trop += len(group) - 1
    return nb_groupes, nb_en_trop


def _import_crm_file_to_db(filepath: Path) -> int:
    """Importe les entreprises d'un fichier CRM dans leads.db — doublons exclus automatiquement."""
    from core.models import Lead, LeadStatus
    from core.crm_filter import load_crm, is_in_crm

    def _int(v):
        try: return int(float(v)) if v is not None else None
        except Exception:
            log.warning("Conversion int echouee pour valeur '%s'", v, exc_info=True)
            return None

    def _float(v):
        try: return float(v) if v is not None else None
        except Exception:
            log.warning("Conversion float echouee pour valeur '%s'", v, exc_info=True)
            return None

    rows = parse_crm_file(filepath)
    if not rows:
        return 0

    # CRM existant, sans le fichier en cours d'import (pour éviter les faux positifs)
    existing_crm = [e for e in load_crm() if e["source_file"] != filepath.name]

    q = LeadQueue(
        db_path=str(ROOT / config.db_path),
        session_id=f"crm_{filepath.stem}",
        session_label=f"CRM — {filepath.name}",
    )
    imported = 0
    for r in rows:
        if existing_crm and is_in_crm(r["name"], r["city"], r["siren"] or "", existing_crm):
            continue
        q.save(Lead(
            company_name=r["name"],
            city=r["city"],
            sector=r["sector"] or "",
            source="crm_import",
            siren=r["siren"] or None,
            website_url=r["website"],
            owner_name=r["owner"],
            owner_role=r["role"],
            phone=r["phone"],
            email=r["email"],
            address=r["address"],
            google_rating=_float(r["rating"]),
            review_count=_int(r["reviews"]),
            legal_form=r["legal"],
            employee_range=r["employee"],
            cms=r["cms"],
            hosting=r["hosting"],
            pagespeed_mobile=_int(r["pagespeed"]),
            seo_score=_int(r["seo_score"]),
            seo_weaknesses=r["seo_weak"].split("|") if r["seo_weak"] else [],
            creation_date=r["creation"],
            etat=r["etat"],
            status=LeadStatus.SCRAPED,
        ))
        imported += 1
    return imported


def _import_to_analyses_db(filepath: Path) -> int:
    """Importe un fichier dans analyses_a2/analyses.db — sans toucher leads.db ni le CRM."""
    from core.models import Lead, LeadStatus

    def _int(v):
        try: return int(float(v)) if v is not None else None
        except Exception:
            log.warning("Conversion int echouee pour valeur '%s' (analyses_db)", v, exc_info=True)
            return None

    def _float(v):
        try: return float(v) if v is not None else None
        except Exception:
            log.warning("Conversion float echouee pour valeur '%s' (analyses_db)", v, exc_info=True)
            return None

    rows = parse_crm_file(filepath)
    if not rows:
        return 0
    ANALYSES_A2_DIR.mkdir(exist_ok=True)
    q = LeadQueue(
        db_path=str(ANALYSES_A2_DB),
        session_id=f"a2_{filepath.stem}",
        session_label=f"Analyse — {filepath.name}",
    )
    # WAL mode : autorise les lectures simultanées pendant que l'agent écrit
    with sqlite3.connect(str(ANALYSES_A2_DB)) as _wc:
        _wc.execute("PRAGMA journal_mode=WAL")
    for r in rows:
        q.save(Lead(
            company_name=r["name"],
            city=r["city"],
            sector=r["sector"] or "",
            source="analyse_a2",
            siren=r["siren"] or None,
            website_url=r["website"],
            owner_name=r["owner"],
            owner_role=r["role"],
            phone=r["phone"],
            email=r["email"],
            address=r["address"],
            google_rating=_float(r["rating"]),
            review_count=_int(r["reviews"]),
            legal_form=r["legal"],
            employee_range=r["employee"],
            cms=r["cms"],
            hosting=r["hosting"],
            pagespeed_mobile=_int(r["pagespeed"]),
            seo_score=_int(r["seo_score"]),
            seo_weaknesses=r["seo_weak"].split("|") if r["seo_weak"] else [],
            creation_date=r["creation"],
            etat=r["etat"],
            status=LeadStatus.SCRAPED,
        ))
    return len(rows)


# Colonnes export
COLS_A1 = {
    "company_name":"Entreprise","city":"Ville","sector":"Secteur","address":"Adresse",
    "google_rating":"Note","review_count":"Avis",
    "owner_name":"Dirigeant","owner_role":"Rôle","phone":"Téléphone",
    "website_url":"Site web","gmb_confirmed":"GMB",
    "siren":"SIREN","siret":"SIRET","tva_intra":"TVA",
    "legal_form":"Forme","etat":"État","employee_range":"Effectif",
    "naf_code":"NAF","naf_label":"Activité","creation_date":"Création",
    "siege_adresse":"Adresse siège","siege_cp":"CP","siege_ville":"Ville siège",
    "nb_etab_ouverts":"Établissements","source":"Source",
}
COLS_A2 = {
    "company_name":"Entreprise","city":"Ville","website_url":"Site web",
    "email":"Email","phone":"Téléphone",
    "cms":"CMS","hosting":"Hébergeur","agence_web":"Agence web",
    "is_https":"HTTPS","pagespeed_mobile":"Speed mobile","pagespeed_desktop":"Speed desktop",
    "copyright_year":"© Année","is_responsive":"Responsive","has_analytics":"Analytics",
    "has_meta_desc":"Meta desc","h1_count":"Nb H1",
    "has_seo_keywords":"Keywords SEO","seo_signals":"Signaux SEO",
    "seo_score":"Score SEO /10","seo_weaknesses":"Faiblesses détectées",
}
COLS_VIEW = {
    "company_name":"Entreprise","city":"Ville","sector":"Secteur",
    "google_rating":"Note","owner_name":"Dirigeant","phone":"Téléphone",
    "email":"Email","website_url":"Site","cms":"CMS","hosting":"Hébergeur",
    "pagespeed_mobile":"Speed mob","has_seo_keywords":"SEO ok",
    "seo_weaknesses":"Faiblesses","employee_range":"Effectif",
    "legal_form":"Forme","siren":"SIREN","etat":"État",
}
COLS_VENDEUR = {
    "company_name":   "Entreprise",
    "city":           "Ville",
    "sector":         "Secteur",
    "owner_name":     "Dirigeant",
    "owner_role":     "Rôle",
    "phone":          "Téléphone",
    "email":          "Email",
    "website_url":    "Site web",
    "google_rating":  "Note Google",
    "review_count":   "Avis",
    "legal_form":     "Forme juridique",
    "employee_range": "Effectif",
    "cms":            "CMS",
    "hosting":        "Hébergeur",
    "pagespeed_mobile": "Vitesse mobile",
    "seo_score":      "Score SEO /10",
    "seo_weaknesses": "Faiblesses SEO",
    "creation_date":  "Création",
    "etat":           "État",
    "siren":          "SIREN",
    "address":        "Adresse",
}


# ── Header ────────────────────────────────────────────────────────────────────
stats = db_stats()
st.markdown("<div style='padding:22px 40px 0'>", unsafe_allow_html=True)
c_logo, c_stats, c_site, c_quit = st.columns([1, 3, 0.5, 0.4])
with c_logo:
    st.markdown(
        '<div style="font-family:Epilogue,sans-serif;font-size:18px;font-weight:900;color:#F0EBE3;letter-spacing:-0.5px">'
        '◈ leads<span style="color:#E87B2A">.</span>engine</div>'
        '<div style="font-size:10px;color:#4A4D58;letter-spacing:2px;text-transform:uppercase;margin-top:3px">'
        'Prospection automatisée'
        f' <span style="color:#2E3240;margin-left:8px">v{(ROOT / "version.txt").read_text(encoding="utf-8").strip() if (ROOT / "version.txt").exists() else "?"}</span>'
        f' <span style="background:{"#E87B2A" if _IS_PRO else "#4A4D58"};color:#0D0E11;font-size:9px;font-weight:800;padding:2px 6px;border-radius:3px;margin-left:6px">{_TIER.upper()}</span>'
        '</div>',
        unsafe_allow_html=True
    )
with c_stats:
    if stats:
        cx1, cx2, cx3, cx4 = st.columns(4)
        cx1.metric("Leads",           stats.get("total", 0))
        cx2.metric("Sessions",        stats.get("sessions", 0))
        cx3.metric("Analysés A2",     stats.get("analyses", 0))
        cx4.metric("Leads qualifiés", stats.get("qualifies", 0))
with c_site:
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    st.link_button("↗ Site web", "https://leadsengine.netlify.app", type="secondary")
with c_quit:
    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    if st.button("✕ Fermer", type="secondary", key="btn_quit_header"):
        # Ferme l'onglet navigateur via JavaScript
        st.markdown(
            '<script>window.open("","_self");window.close();</script>',
            unsafe_allow_html=True,
        )
        import signal
        if getattr(sys, "frozen", False):
            _os.system('taskkill /F /IM LeadsEngine.exe >nul 2>&1')
        else:
            _os.kill(_os.getpid(), signal.SIGTERM)
st.markdown("</div>", unsafe_allow_html=True)
st.markdown("<div style='height:1px;background:#1E2028;margin:16px 0 0'></div>", unsafe_allow_html=True)


# ── Bannière setup si configuration incomplète ────────────────────────────────
if not config.serpapi_key:
    st.warning("⚙️ Configuration incomplète — rends-toi dans l'onglet **Configuration** pour entrer ta clé SERPAPI_KEY avant de scraper.", icon=None)

# ── Auto-update ───────────────────────────────────────────────────────────────
if "update_checked" not in st.session_state:
    try:
        from core.updater import check_update, get_local_version
        st.session_state["update_info"] = check_update(ROOT)
        st.session_state["app_version"] = get_local_version(ROOT)
    except Exception:
        st.session_state["update_info"] = None
        st.session_state["app_version"] = "?"
    st.session_state["update_checked"] = True

_upd = st.session_state.get("update_info")
if _upd:
    _size_mb = _upd.get("size", 0) / (1024 * 1024)
    st.markdown(
        f'<div style="background:#1A1200;border:1px solid #3A2E00;border-left:3px solid #E8C32A;'
        f'border-radius:8px;padding:12px 18px;margin-bottom:12px;display:flex;'
        f'align-items:center;justify-content:space-between">'
        f'<span style="color:#F5D87A;font-size:13px;font-weight:600">'
        f'Nouvelle version {_upd["version"]} disponible ({_size_mb:.0f} Mo)</span>'
        f'</div>',
        unsafe_allow_html=True,
    )
    if st.button("Mettre à jour", type="primary", key="btn_update"):
        _prog = st.progress(0, text="Téléchargement en cours…")
        _ok = download_and_install(
            ROOT, _upd,
            progress_callback=lambda p: _prog.progress(
                min(p, 1.0), text=f"Téléchargement… {int(p * 100)} %"),
        )
        if _ok:
            _prog.progress(1.0, text="Téléchargement terminé !")
            st.success("Mise à jour téléchargée — fermeture et installation…")
            time.sleep(1)
            launch_update_and_quit(ROOT)
        else:
            st.error("Erreur lors de la mise à jour. Consulte errors.log.")

# ── Navigation ────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5, tab7, tab6 = st.tabs([
    "🚀  Scraping",
    "🔍  Analyse sites",
    "📋  Mes recherches",
    "🗂  CRM",
    "🌐  Tous les leads",
    "📇  Fiches Leads",
    "⚙️  Configuration",
])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — Scraping
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    st.title("Scraping")
    st.markdown('<p style="color:#4A4D58;font-size:13px;margin-bottom:24px">Collecte les entreprises, dirigeants et données du registre national.</p>', unsafe_allow_html=True)

    col1, col2 = st.columns([3, 2], gap="large")
    with col1:
        st.markdown('<div class="section-lbl">Cible</div>', unsafe_allow_html=True)
        secteurs    = st.text_input("Secteurs", value="plombier,électricien", key="sec1")
        ville = st.text_input("Ville(s)", value="Aix-en-Provence", key="vil1",
                              disabled=st.session_state.get("france_entiere", False))
        _france_entiere = st.toggle("France entière", value=False, key="france_entiere")
        if _france_entiere:
            ville = ""
            st.caption("Le scraper parcourra automatiquement les principales villes de France.")
        nom_session = st.text_input("Nom de la recherche", placeholder="Plombiers Aix Juin 2025", key="nom1")
    with col2:
        st.markdown('<div class="section-lbl">Paramètres</div>', unsafe_allow_html=True)
        max_r_total = st.slider("Nombre de leads souhaités", 10, 500, 50, 10)
        use_maps     = st.toggle("Google Maps", value=True)
        use_registre = st.toggle("Registre National", value=True)
        _site_filter = st.radio("Filtre site web", ["Tous", "Avec site uniquement", "Sans site uniquement"], horizontal=True, key="site_filter")

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    if st.button("Lancer le scraping →", type="primary", use_container_width=True, key="btn_scrap"):
        if not secteurs.strip():
            st.error("Remplis au moins un secteur.")
        elif not _france_entiere and not ville.strip():
            st.error("Remplis une ville ou active 'France entière'.")
        elif use_maps and not config.serpapi_key:
            st.error("SERPAPI_KEY manquante dans ton fichier .env")
        else:
            label      = nom_session.strip() or datetime.now().strftime("%Y%m%d_%H%M%S")
            session_id = re.sub(r"[^\w\-]", "_", label)
            if _france_entiere:
                _sectors_list = [s.strip() for s in secteurs.split(",") if s.strip()]
                queries = [(s, "__france__") for s in _sectors_list]
                _per_query = max_r_total
            else:
                queries = [(s.strip(), v.strip()) for s in secteurs.split(",") for v in ville.split(",")]
                _per_query = max(5, -(-max_r_total // len(queries)))  # ceil division
            prog = st.progress(10, text="Connexion...")
            rh, eh = [], []

            agents_ref = []

            def _sc():
                try:
                    loop = asyncio.new_event_loop()
                    asyncio.set_event_loop(loop)
                    q = LeadQueue(str(ROOT / config.db_path), session_id=session_id, session_label=label)
                    a = ScraperAgent(q)
                    agents_ref.append(a)
                    _wf = "with_site" if _site_filter == "Avec site uniquement" else "without_site" if _site_filter == "Sans site uniquement" else "all"
                    rh.append(loop.run_until_complete(a.run(
                        queries=queries, use_maps=use_maps,
                        use_registre=use_registre,
                        max_per_query=_per_query,
                        max_total=max_r_total,
                        website_filter=_wf,
                    )))
                    loop.close()
                except Exception as e:
                    log.error("Thread scraping (Agent 1) erreur", exc_info=True)
                    eh.append(e)

            t = threading.Thread(target=_sc, daemon=True)
            t.start()
            prog.progress(40, text="Scraping en cours...")
            _t0 = time.time()
            _chron1 = st.empty()
            while t.is_alive():
                time.sleep(1)
                _e = int(time.time() - _t0)
                _mm, _ss = divmod(_e, 60)
                _chron1.markdown(
                    f'<span style="font-family:IBM Plex Mono,monospace;font-size:12px;color:#4A4D58">⏱ {_mm:02d}:{_ss:02d}</span>',
                    unsafe_allow_html=True
                )
            t.join(timeout=600)
            prog.progress(100, text="Terminé ✓")
            _ef = int(time.time() - _t0)
            _mmf, _ssf = divmod(_ef, 60)
            _chron1.markdown(
                f'<span style="font-family:IBM Plex Mono,monospace;font-size:12px;color:#E87B2A">⏱ terminé en {_mmf:02d}:{_ssf:02d}</span>',
                unsafe_allow_html=True
            )

            if eh:
                st.error(f"Erreur : {eh[0]}")
            elif rh:
                leads = rh[0]
                wo = sum(1 for l in leads if l.owner_name)
                ws = sum(1 for l in leads if l.website_url)
                st.success(f"{len(leads)} leads collectés — session « {label} »")
                cx1, cx2, cx3 = st.columns(3)
                cx1.metric("Total", len(leads))
                cx2.metric("Avec site", ws)
                cx3.metric("Avec dirigeant", wo)

                # Afficher les villes explorées si extension automatique
                if agents_ref and getattr(agents_ref[0], "expanded_cities", None):
                    for query_key, cities_list in agents_ref[0].expanded_cities.items():
                        ville_origin = cities_list[0]
                        villes_ext = cities_list[1:]
                        st.info(
                            f"🗺️ **Extension automatique** pour *{query_key}* :\n"
                            f"Ville demandée **{ville_origin}** saturée → "
                            f"recherche étendue à **{len(villes_ext)}** ville(s) voisine(s) :\n"
                            f"{', '.join(villes_ext)}"
                        )

                st.info("→ Lance l'**Analyse sites** pour obtenir CMS, hébergeur, vitesse et SEO keywords.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — Agent 2
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.title("Analyse des sites web")
    st.markdown('<p style="color:#4A4D58;font-size:13px;margin-bottom:24px">Extrait CMS, hébergeur, vitesse, email, agence et signaux SEO. 100% gratuit.</p>', unsafe_allow_html=True)

    db = ROOT / config.db_path
    if not db.exists():
        st.info("Lance d'abord un scraping.")
    else:
        queue2   = LeadQueue(str(ROOT / config.db_path))
        sessions = queue2.list_sessions()
        if not sessions:
            st.info("Aucune session trouvée.")
        else:
            session_map2 = {fmt_session(s): s["session_id"] for s in sessions}
            choix2   = st.selectbox("Session à analyser", list(session_map2.keys()), key="sess2")
            sid2     = session_map2[choix2]
            s2       = next(x for x in sessions if x["session_id"] == sid2)

            conn = sqlite3.connect(db)
            n_site = conn.execute("SELECT COUNT(*) FROM leads WHERE session_id=? AND website_url IS NOT NULL", (sid2,)).fetchone()[0]
            n_done = conn.execute("SELECT COUNT(*) FROM leads WHERE session_id=? AND status='extracted'", (sid2,)).fetchone()[0]
            conn.close()

            st.divider()
            cx1, cx2, cx3, cx4 = st.columns(4)
            cx1.metric("Total",     s2["total"])
            cx2.metric("Avec site", n_site)
            cx3.metric("Analysés",  n_done)
            cx4.metric("Restants",  max(0, n_site - n_done))

            st.divider()
            col1, col2 = st.columns([2, 1])
            with col1:
                delay    = st.slider("Délai entre visites (sec)", 1.0, 6.0, 2.0, 0.5)
                only_new = st.checkbox("Seulement les non analysés", value=True)

            with col2:
                reste = max(0, n_site - n_done) if only_new else n_site
                _extra_time = 0
                st.markdown(
                    f'<div style="margin-top:24px;background:#13151A;border:1px solid #1E2028;border-radius:8px;padding:14px 18px">'
                    f'<div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:1.2px;color:#4A4D58;margin-bottom:6px">Temps estimé</div>'
                    f'<div style="font-family:IBM Plex Mono,monospace;font-size:22px;color:#E87B2A">~{reste*(delay+40)/60:.0f} min</div>'
                    f'<div style="font-size:11px;color:#4A4D58;margin-top:2px">{reste} sites</div></div>',
                    unsafe_allow_html=True
                )

            with st.expander("Modules d'extraction", expanded=False):
                st.caption("Sélectionne les données à extraire. Désactiver un module accélère l'analyse.")
                _mc1, _mc2, _mc3 = st.columns(3)
                with _mc1:
                    _mod_cms       = st.checkbox("CMS & hébergeur",    value=True, key="mod_cms")
                    _mod_seo       = st.checkbox("SEO & structure",    value=True, key="mod_seo")
                    _mod_contacts  = st.checkbox("Email & téléphone",  value=True, key="mod_contacts")
                with _mc2:
                    _mod_pagespeed = st.checkbox("PageSpeed",          value=True, key="mod_pagespeed")
                    _mod_social    = st.checkbox("Réseaux sociaux",    value=True, key="mod_social", disabled=not _IS_PRO, help="Pro uniquement" if not _IS_PRO else None)
                    _mod_agence    = st.checkbox("Agence web",         value=True, key="mod_agence")
                with _mc3:
                    _mod_domain    = st.checkbox("Age du domaine",     value=True, key="mod_domain")
                    _mod_gads      = st.checkbox("Google Ads",         value=True, key="mod_gads", disabled=not _IS_PRO, help="Pro uniquement" if not _IS_PRO else None)

                _a2_modules = {
                    "cms": _mod_cms, "seo": _mod_seo, "contacts": _mod_contacts,
                    "pagespeed": _mod_pagespeed, "social": _mod_social, "agence": _mod_agence,
                    "domain_age": _mod_domain, "google_ads": _mod_gads,
                }

            if st.button("Lancer l'analyse →", type="primary", use_container_width=True, key="btn_a2"):
                conn = sqlite3.connect(db)
                conn.row_factory = sqlite3.Row
                if only_new:
                    sql = "SELECT * FROM leads WHERE session_id=? AND website_url IS NOT NULL AND status='scraped'"
                    rows = conn.execute(sql, (sid2,)).fetchall()
                else:
                    # Relance complète : remettre tous les leads avec site en 'scraped'
                    sql_all = "SELECT * FROM leads WHERE session_id=? AND website_url IS NOT NULL"
                    rows = conn.execute(sql_all, (sid2,)).fetchall()
                    ids_reset = [r["id"] for r in rows]
                    if ids_reset:
                        conn.executemany("UPDATE leads SET status='scraped' WHERE id=?", [(i,) for i in ids_reset])
                        conn.commit()
                        rows = conn.execute(sql_all, (sid2,)).fetchall()
                conn.close()
                leads_to_do = [dict(r) for r in rows]

                if not leads_to_do:
                    if n_site == 0:
                        st.warning("Cette session ne contient aucun lead avec un site web — rien à analyser.")
                    else:
                        st.warning("Aucun lead à analyser.")
                else:
                    prog2 = st.progress(0, text="Démarrage...")
                    rh2, eh2 = [], []

                    def _a2():
                        try:
                            rh2.append(ExtractorAgent(queue2).run(leads_to_do, delay=delay, modules=_a2_modules))
                        except Exception as e:
                            log.error("Thread Agent 2 erreur", exc_info=True)
                            eh2.append(e)

                    t2 = threading.Thread(target=_a2, daemon=True)
                    t2.start()
                    total2 = len(leads_to_do)
                    _t0_a2 = time.time()
                    _chron2 = st.empty()
                    while t2.is_alive():
                        time.sleep(2)
                        _e2 = int(time.time() - _t0_a2)
                        _mm2, _ss2 = divmod(_e2, 60)
                        _chron2.markdown(
                            f'<span style="font-family:IBM Plex Mono,monospace;font-size:12px;color:#4A4D58">⏱ {_mm2:02d}:{_ss2:02d}</span>',
                            unsafe_allow_html=True
                        )
                        try:
                            c3 = sqlite3.connect(db)
                            done2 = c3.execute(
                                "SELECT COUNT(*) FROM leads WHERE session_id=? AND status IN ('extracted','skipped','error')",
                                (sid2,)
                            ).fetchone()[0] - n_done
                            c3.close()
                            pct = min(int(done2 / total2 * 100), 99) if total2 else 0
                            prog2.progress(pct, text=f"{done2}/{total2} sites traités...")
                        except Exception:
                            log.warning("Polling progression Agent 2 echoue", exc_info=True)
                    t2.join(timeout=7200)
                    prog2.progress(100, text="Terminé ✓")
                    _ef2 = int(time.time() - _t0_a2)
                    _mmf2, _ssf2 = divmod(_ef2, 60)
                    _chron2.markdown(
                        f'<span style="font-family:IBM Plex Mono,monospace;font-size:12px;color:#E87B2A">⏱ terminé en {_mmf2:02d}:{_ssf2:02d}</span>',
                        unsafe_allow_html=True
                    )

                    if eh2:
                        st.error(f"Erreur : {eh2[0]}")
                    elif rh2:
                        r2 = rh2[0]
                        st.success(f"Analyse terminée — {r2['success']} analysés · {r2['skipped']} ignorés · {r2['errors']} erreurs")
                        st.info("→ Va dans **Mes recherches** pour voir et exporter les résultats.")

    # ── Analyser un fichier CRM ──────────────────────────────────────────────
    st.divider()
    with st.expander("◈  Analyser un fichier CRM"):
        st.markdown(
            '<p style="color:#4A4D58;font-size:13px;margin-bottom:4px">Importe un export client (Excel / CSV) et lance l\'Agent 2 dessus.</p>'
            '<p style="color:#4A4D58;font-size:12px;margin-bottom:16px">Les résultats sont stockés dans <code>analyses_a2/</code> — sans toucher à leads.db ni au CRM.</p>',
            unsafe_allow_html=True
        )

        try:
            crm_files = sorted(list(CRM_DIR.glob("*.xlsx")) + list(CRM_DIR.glob("*.csv"))) if CRM_DIR.exists() else []
        except Exception:
            log.warning("Listing fichiers CRM echoue", exc_info=True)
            crm_files = []

        src = st.radio("Source", ["Fichier déjà dans le CRM", "Importer un nouveau fichier"], horizontal=True, key="crm_a2_src")

        pending_path = None
        if src == "Fichier déjà dans le CRM":
            if not crm_files:
                st.info("Aucun fichier CRM — ajoutes-en dans l'onglet CRM.")
            else:
                file_map_a2 = {f.name: f for f in crm_files}
                chosen = st.selectbox("Fichier", list(file_map_a2.keys()), key="crm_a2_file")
                pending_path = str(file_map_a2[chosen])
        else:
            up = st.file_uploader("Fichier à analyser", type=["xlsx", "csv"], key="crm_a2_up", label_visibility="collapsed")
            if up:
                ANALYSES_A2_DIR.mkdir(exist_ok=True)
                dest_up = ANALYSES_A2_DIR / up.name
                dest_up.write_bytes(up.read())
                pending_path = str(dest_up)

        # Bouton de chargement explicite — rien ne s'exécute automatiquement
        if pending_path and st.button("Charger ce fichier →", key="btn_crm_load", type="secondary"):
            st.session_state["crm_a2_loaded"] = pending_path

        sid_crm = None
        if "crm_a2_loaded" in st.session_state:
            loaded_path = Path(st.session_state["crm_a2_loaded"])
            if loaded_path.exists():
                sid_crm = f"a2_{loaded_path.stem}"
                _import_key = f"_a2_imported_{sid_crm}"
                if not st.session_state.get(_import_key):
                    try:
                        _import_to_analyses_db(loaded_path)
                        st.session_state[_import_key] = True
                    except Exception as e:
                        log.error("Import fichier CRM dans analyses echoue", exc_info=True)
                        st.error(f"Erreur import : {e}")
                        sid_crm = None

        if sid_crm:
            queue_crm = LeadQueue(str(ANALYSES_A2_DB), session_id=sid_crm)
            conn_c = sqlite3.connect(ANALYSES_A2_DB)
            n_total_c = conn_c.execute("SELECT COUNT(*) FROM leads WHERE session_id=?", (sid_crm,)).fetchone()[0]
            n_site_c  = conn_c.execute("SELECT COUNT(*) FROM leads WHERE session_id=? AND website_url IS NOT NULL", (sid_crm,)).fetchone()[0]
            n_done_c  = conn_c.execute("SELECT COUNT(*) FROM leads WHERE session_id=? AND status IN ('extracted','skipped','error')", (sid_crm,)).fetchone()[0]
            conn_c.close()

            st.divider()
            cc1, cc2, cc3, cc4 = st.columns(4)
            cc1.metric("Total",        n_total_c)
            cc2.metric("Avec site",    n_site_c)
            cc3.metric("Analysés",     n_done_c)
            cc4.metric("Restants",     max(0, n_site_c - n_done_c))

            # ── Enrichissement sites web ─────────────────────────────────────
            n_sans_site = n_total_c - n_site_c
            if n_sans_site > 0:
                st.divider()
                enr_col1, enr_col2 = st.columns([3, 1])
                with enr_col1:
                    st.caption(f"**{n_sans_site} entreprise(s) sans site web** — recherche automatique via Google Maps")
                with enr_col2:
                    if not config.serpapi_key:
                        st.warning("Clé SerpAPI manquante", icon=None)
                    elif st.button("Trouver les sites →", key="btn_enrich_web", type="secondary", use_container_width=True):
                        delay_enr = 1.5
                        enr_rh, enr_eh = [], []
                        enr_thread = threading.Thread(
                            target=_enrich_websites,
                            args=(sid_crm, str(ANALYSES_A2_DB), delay_enr, enr_rh, enr_eh),
                            daemon=True
                        )
                        enr_thread.start()
                        enr_prog = st.progress(0, text="Recherche en cours…")
                        while enr_thread.is_alive():
                            time.sleep(2)
                            try:
                                _nc = sqlite3.connect(str(ANALYSES_A2_DB))
                                _found_so_far = _nc.execute(
                                    "SELECT COUNT(*) FROM leads WHERE session_id=? AND website_url IS NOT NULL",
                                    (sid_crm,)
                                ).fetchone()[0] - n_site_c
                                _nc.close()
                                pct = min(int(_found_so_far / n_sans_site * 100), 99) if n_sans_site else 0
                                enr_prog.progress(pct, text=f"{_found_so_far}/{n_sans_site} sites trouvés…")
                            except Exception:
                                log.warning("Polling enrichissement sites echoue", exc_info=True)
                        enr_thread.join(timeout=3600)
                        enr_prog.progress(100, text="Terminé ✓")
                        if enr_eh:
                            st.error(f"Erreur : {enr_eh[0]}")
                        elif enr_rh:
                            found_total, searched_total = enr_rh[0]
                            st.success(f"{found_total} site(s) trouvé(s) sur {searched_total} entreprise(s) recherchées.")
                        st.rerun()

            col_ca, col_cb = st.columns([2, 1])
            with col_ca:
                delay_c    = st.slider("Délai entre visites (sec)", 1.0, 6.0, 2.0, 0.5, key="delay_crm")
                only_new_c = st.checkbox("Seulement les non analysés", value=True, key="only_crm")
            with col_cb:
                reste_c = max(0, n_site_c - n_done_c) if only_new_c else n_site_c
                st.markdown(
                    f'<div style="margin-top:24px;background:#13151A;border:1px solid #1E2028;border-radius:8px;padding:14px 18px">'
                    f'<div style="font-size:10px;font-weight:700;text-transform:uppercase;letter-spacing:1.2px;color:#4A4D58;margin-bottom:6px">Temps estimé</div>'
                    f'<div style="font-family:IBM Plex Mono,monospace;font-size:22px;color:#E87B2A">~{reste_c*(delay_c+20)/60:.0f} min</div>'
                    f'<div style="font-size:11px;color:#4A4D58;margin-top:2px">{reste_c} sites</div></div>',
                    unsafe_allow_html=True
                )

            if st.button("Analyser →", type="primary", use_container_width=True, key="btn_a2_crm"):
                conn_c2 = sqlite3.connect(ANALYSES_A2_DB)
                conn_c2.row_factory = sqlite3.Row
                sql_c = ("SELECT * FROM leads WHERE session_id=? AND website_url IS NOT NULL AND status='scraped'"
                         if only_new_c else "SELECT * FROM leads WHERE session_id=? AND website_url IS NOT NULL")
                rows_c = conn_c2.execute(sql_c, (sid_crm,)).fetchall()
                # Si relance complète, remettre tous ces leads en 'scraped' pour un suivi correct
                if not only_new_c:
                    ids_to_reset = [r["id"] for r in rows_c]
                    if ids_to_reset:
                        conn_c2.executemany(
                            "UPDATE leads SET status='scraped' WHERE id=?",
                            [(i,) for i in ids_to_reset]
                        )
                        conn_c2.commit()
                        # Recharger avec le statut réinitialisé
                        rows_c = conn_c2.execute(sql_c, (sid_crm,)).fetchall()
                conn_c2.close()
                leads_c = [dict(r) for r in rows_c]

                if not leads_c:
                    if n_site_c == 0:
                        st.warning("Ce fichier ne contient aucun site web reconnu — rien à analyser.")
                    else:
                        st.warning("Aucun lead à analyser.")
                else:
                    prog_c = st.progress(0, text="Démarrage...")
                    rh_c, eh_c = [], []

                    def _a2_crm():
                        try:
                            rh_c.append(ExtractorAgent(queue_crm).run(leads_c, delay=delay_c))
                        except Exception as e:
                            log.error("Thread Agent 2 CRM erreur", exc_info=True)
                            eh_c.append(e)

                    tc = threading.Thread(target=_a2_crm, daemon=True)
                    tc.start()
                    total_c = len(leads_c)
                    _t0_ac = time.time()
                    _chron_c = st.empty()
                    while tc.is_alive():
                        time.sleep(2)
                        _ec = int(time.time() - _t0_ac)
                        _mmc, _ssc = divmod(_ec, 60)
                        _chron_c.markdown(
                            f'<span style="font-family:IBM Plex Mono,monospace;font-size:12px;color:#4A4D58">⏱ {_mmc:02d}:{_ssc:02d}</span>',
                            unsafe_allow_html=True
                        )
                        try:
                            cc = sqlite3.connect(ANALYSES_A2_DB)
                            done_c = cc.execute(
                                "SELECT COUNT(*) FROM leads WHERE session_id=? AND status IN ('extracted','skipped','error')",
                                (sid_crm,)
                            ).fetchone()[0] - n_done_c
                            cc.close()
                            pct_c = min(int(done_c / total_c * 100), 99) if total_c else 0
                            prog_c.progress(pct_c, text=f"{done_c}/{total_c} sites traités...")
                        except Exception:
                            log.warning("Polling progression Agent 2 CRM echoue", exc_info=True)
                    tc.join(timeout=7200)
                    prog_c.progress(100, text="Terminé ✓")
                    _efc = int(time.time() - _t0_ac)
                    _mmfc, _ssfc = divmod(_efc, 60)
                    _chron_c.markdown(
                        f'<span style="font-family:IBM Plex Mono,monospace;font-size:12px;color:#E87B2A">⏱ terminé en {_mmfc:02d}:{_ssfc:02d}</span>',
                        unsafe_allow_html=True
                    )

                    if eh_c:
                        st.error(f"Erreur : {eh_c[0]}")
                    elif rh_c:
                        rc = rh_c[0]
                        st.success(f"Analyse terminée — {rc['success']} analysés · {rc['skipped']} ignorés · {rc['errors']} erreurs")
                        # Export automatique dans analyses_a2/
                        try:
                            conn_ex = sqlite3.connect(ANALYSES_A2_DB)
                            df_ex = pd.read_sql(
                                "SELECT * FROM leads WHERE session_id=?", conn_ex, params=(sid_crm,)
                            )
                            conn_ex.close()
                            xl_ex = to_excel_combined(df_ex, COLS_A1, COLS_A2)
                            fname_ex = f"{loaded_path.stem}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                            out_path = ANALYSES_A2_DIR / fname_ex
                            out_path.write_bytes(xl_ex)
                            st.download_button(
                                f"Télécharger les résultats ({len(df_ex)} leads)",
                                data=xl_ex, file_name=fname_ex,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True, type="primary",
                            )
                            st.caption(f"Enregistré dans analyses_a2/{fname_ex}")
                        except Exception as e:
                            log.error("Export Excel apres analyse CRM echoue", exc_info=True)
                            st.error(f"Erreur export : {e}")



# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — Mes recherches
# ══════════════════════════════════════════════════════════════════════════════
with tab3:
    st.title("Mes recherches")
    st.markdown('<p style="color:#4A4D58;font-size:13px;margin-bottom:24px">Chaque session est isolée. Filtre, explore et exporte.</p>', unsafe_allow_html=True)

    db = ROOT / config.db_path
    if not db.exists():
        st.info("Aucune recherche effectuée.")
    else:
        queue3   = LeadQueue(str(ROOT / config.db_path))
        sessions = queue3.list_sessions()
        if not sessions:
            st.info("Aucune session.")
        else:
            vue_globale = st.toggle("Toute la base", key="vue_globale3", help="Affiche tous les leads toutes sessions confondues")

            if vue_globale:
                df   = load_df()
                sid3 = "base_complete"
            else:
                session_map3 = {fmt_session(s): s["session_id"] for s in sessions}
                choix3   = st.selectbox("Session", list(session_map3.keys()), key="sess3")
                sid3     = session_map3[choix3]
                df = load_df(sid3)

            a2_ok    = has_agent2(df)
            nb_qual  = int(df.apply(is_qualifie, axis=1).sum()) if not df.empty else 0
            nb_email = int(df["email"].notna().sum())  if "email" in df.columns else 0
            nb_cms   = int(df["cms"].notna().sum())    if "cms"   in df.columns else 0
            nb_seo   = int(df["has_seo_keywords"].notna().sum()) if "has_seo_keywords" in df.columns else 0

            st.divider()
            cx1, cx2, cx3, cx4, cx5, cx6 = st.columns(6)
            cx1.metric("Total",           len(df))
            cx2.metric("Avec site",       int(df["website_url"].notna().sum()) if "website_url" in df.columns else 0)
            cx3.metric("Dirigeant",       int(df["owner_name"].notna().sum())  if "owner_name"  in df.columns else 0)
            cx4.metric("Email trouvé",    nb_email)
            cx5.metric("CMS détecté",     nb_cms)
            cx6.metric("Leads qualifiés", nb_qual)

            if a2_ok:
                st.markdown(
                    f'<div class="a2-banner">'
                    f'<div class="t">◈ Données Agent 2 disponibles</div>'
                    f'<div class="d">{nb_cms} CMS · {nb_seo} scores SEO · {nb_email} emails</div>'
                    f'</div>',
                    unsafe_allow_html=True
                )

            st.divider()

            if not df.empty:
                search = st.text_input("🔍  Rechercher par nom", placeholder="ex: Dupont, Boulangerie…", key="search3", label_visibility="collapsed")

                col_f1, col_f2, col_f3, col_f4 = st.columns(4)
                with col_f1:
                    f_qual = st.toggle(
                        "Leads qualifiés uniquement", key="fq3",
                        help="Nom + adresse + tél + ville + secteur + dirigeant + CMS + hébergeur + vitesse + SEO"
                    )
                with col_f2:
                    f_email = st.checkbox("Avec email", key="fe3")
                with col_f3:
                    f_site  = st.checkbox("Avec site web", key="fs3")
                with col_f4:
                    f_a2v = st.checkbox("Agent 2 analysé", key="fa3") if a2_ok else False

                df_show = df.copy()
                if search:          df_show = df_show[df_show["company_name"].str.contains(search, case=False, na=False)]
                if f_qual:          df_show = df_show[df_show.apply(is_qualifie, axis=1)]
                if f_email:         df_show = df_show[df_show["email"].notna()]
                if f_site:          df_show = df_show[df_show["website_url"].notna()]
                if a2_ok and f_a2v: df_show = df_show[df_show["cms"].notna()]

                st.caption(f"{len(df_show)} leads affichés")
                cols_view = [c for c in COLS_VIEW if c in df_show.columns]
                st.dataframe(
                    df_show[cols_view].rename(columns=COLS_VIEW),
                    use_container_width=True, hide_index=True, height=400
                )

                # Export
                st.markdown('<div class="export-box"><div class="export-lbl">Export Excel</div>', unsafe_allow_html=True)

                _nom_raw = st.text_input(
                    "Nom du fichier", placeholder=f"ex : Plombiers Marseille Juin 2025",
                    key="export_nom3", label_visibility="visible"
                )
                _nom_base = re.sub(r'[\\/*?:"<>|]', "", _nom_raw.strip()) or sid3

                df_qual_rows = df[df.apply(is_qualifie, axis=1)]

                ce1, ce2, ce3, ce4 = st.columns(4)
                with ce1:
                    # Toutes les colonnes A1+A2, vides si non renseignées (permet la réimportation pour analyse)
                    _all_cols_a1 = {**COLS_A1, **{k: v for k, v in COLS_A2.items() if k not in COLS_A1}}
                    _df_a1_full  = df.copy()
                    for _c in _all_cols_a1:
                        if _c not in _df_a1_full.columns:
                            _df_a1_full[_c] = None
                    _df_a1_full = _df_a1_full[[c for c in _all_cols_a1]].rename(columns=_all_cols_a1)
                    xl1 = to_excel_a1(_df_a1_full)
                    st.download_button(
                        "📄  Agent 1 — Registre national",
                        data=xl1, file_name=f"{_nom_base}_agent1.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True, type="primary",
                        help="Nom, dirigeant, SIREN, SIRET, TVA, forme juridique, effectif, adresse siège..."
                    )
                with ce2:
                    if a2_ok:
                        xl2 = to_excel_combined(df, COLS_A1, COLS_A2)
                        st.download_button(
                            "◈  Agent 1 + Agent 2 (tout en un)",
                            data=xl2, file_name=f"{_nom_base}_complet.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            help="Toutes les colonnes Agent 1 + Agent 2 fusionnées sur un seul onglet"
                        )
                    else:
                        st.markdown('<div style="color:#4A4D58;font-size:11px;padding-top:8px">Lance l\'Agent 2 pour débloquer</div>', unsafe_allow_html=True)
                with ce3:
                    if not df_qual_rows.empty:
                        xl3 = to_excel_qualifies(df_qual_rows, COLS_A1, COLS_A2)
                        st.download_button(
                            f"✓  Leads qualifiés ({len(df_qual_rows)})",
                            data=xl3, file_name=f"{_nom_base}_qualifies.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            help="Leads avec les 10 critères — colonnes A1 + A2 fusionnées"
                        )
                    else:
                        st.markdown('<div style="color:#4A4D58;font-size:11px;padding-top:8px">Aucun lead qualifié pour l\'instant</div>', unsafe_allow_html=True)
                with ce4:
                    xl_v = to_excel_vendeur(df_show)
                    st.download_button(
                        f"🎯  Vue vendeur ({len(df_show)})",
                        data=xl_v, file_name=f"{_nom_base}_vendeur.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        help="Colonnes essentielles pour un commercial — mise en forme incluse"
                    )

                # Export CRM
                _crm_choice = load_user_config().get("crm", "default")
                _all_crm_keys = list(CRM_MAPPINGS.keys())
                _default_idx = _all_crm_keys.index(_crm_choice) if _crm_choice in _all_crm_keys else 0

                st.markdown('<div style="margin-top:12px"><div class="export-lbl">Export CRM</div>', unsafe_allow_html=True)
                _exp_crm = st.selectbox(
                    "CRM", options=_all_crm_keys,
                    index=_default_idx,
                    format_func=lambda x: CRM_MAPPINGS[x]["label"] + (" — Connecté" if is_connected(x) else ""),
                    key=f"exp_crm_{sid3}", label_visibility="collapsed",
                )
                _exp_label = CRM_MAPPINGS[_exp_crm]["label"]

                _exp_c1, _exp_c2 = st.columns(2)
                with _exp_c1:
                    _crm_csv = export_crm_csv(df, _exp_crm)
                    st.download_button(
                        f"Télécharger CSV — {_exp_label}",
                        data=_crm_csv,
                        file_name=f"{_nom_base}_{_exp_crm}.csv",
                        mime="text/csv",
                        use_container_width=True, type="secondary",
                        help=f"CSV prêt à importer manuellement dans {_exp_label}",
                        key=f"crm_csv_{sid3}",
                    )
                with _exp_c2:
                    if is_connected(_exp_crm):
                        if st.button(
                            f"Envoyer {len(df)} leads vers {_exp_label}",
                            type="primary", key=f"push_btn_{sid3}", use_container_width=True,
                        ):
                            _push_bar = st.progress(0, text="Envoi en cours…")
                            _p_ok, _p_fail, _p_errors = push_leads(
                                _exp_crm, df,
                                progress_cb=lambda p: _push_bar.progress(p, text=f"Envoi… {p*100:.0f}%"),
                            )
                            _push_bar.progress(1.0, text="Terminé")
                            _skipped_msg = [e for e in _p_errors if "ignorés" in e]
                            _real_errors = [e for e in _p_errors if "ignorés" not in e]
                            if _p_ok > 0 and _skipped_msg:
                                st.success(f"{_p_ok} leads envoyés vers {_exp_label} — {_skipped_msg[0]}")
                            elif _p_ok > 0:
                                st.success(f"{_p_ok} leads envoyés vers {_exp_label}.")
                            elif _skipped_msg:
                                st.info(f"Aucun nouveau lead à envoyer — {_skipped_msg[0]}")
                            if _p_fail > 0:
                                st.warning(f"{_p_fail} leads en erreur.")
                            if _real_errors:
                                with st.expander("Détails des erreurs"):
                                    for _err in _real_errors[:20]:
                                        st.text(_err)
                    else:
                        st.markdown(
                            f'<div style="text-align:center;padding:8px;color:#4A4D58;font-size:11px">'
                            f'{_exp_label} non connecté<br>Configure-le dans l\'onglet Configuration</div>',
                            unsafe_allow_html=True,
                        )
                st.markdown('</div>', unsafe_allow_html=True)

                st.markdown('</div>', unsafe_allow_html=True)

                st.divider()

                if not vue_globale:
                    with st.expander("Supprimer cette session"):
                        s3 = next(x for x in sessions if x["session_id"] == sid3)
                        st.warning(f"Supprimer définitivement « {s3.get('session_label', sid3)} » ({s3['total']} leads) ?")
                        if st.button("Confirmer la suppression", type="secondary", key="del3"):
                            queue3.delete_session(sid3)
                            st.success("Session supprimée.")
                            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — CRM
# ══════════════════════════════════════════════════════════════════════════════
with tab4:
    if not _IS_PRO:
        st.markdown(
            '<div style="text-align:center;padding:80px 20px">'
            '<div style="font-size:48px;margin-bottom:16px">🔒</div>'
            '<div style="font-size:18px;font-weight:700;color:#F0EBE3;margin-bottom:8px">Fonctionnalité Pro</div>'
            '<div style="font-size:13px;color:#4A4D58">L\'onglet CRM est réservé à la licence Pro.</div>'
            '<a href="https://leadsengine.netlify.app" target="_blank" '
            'style="display:inline-block;margin-top:20px;background:#E87B2A;color:#0D0E11;padding:8px 20px;'
            'border-radius:6px;font-weight:700;font-size:13px;text-decoration:none">Passer en Pro</a>'
            '</div>',
            unsafe_allow_html=True,
        )
    if _IS_PRO:
      try:
        stats_crm = crm_stats()

        st.title("Fichiers CRM")
        st.caption("Les entreprises présentes dans ces fichiers sont exclues des prochains scrapings.")

        m1, m2 = st.columns(2)
        m1.metric("Fichiers chargés",    stats_crm["fichiers"])
        m2.metric("Entreprises connues", stats_crm["entreprises"])

        st.divider()
        st.subheader("Ajouter un fichier")
        uploaded = st.file_uploader(
            "Ajouter", type=["xlsx", "csv"],
            accept_multiple_files=True, key="crm_upload",
            label_visibility="collapsed",
        )
        if uploaded:
            CRM_DIR.mkdir(exist_ok=True)
            n_imp = 0
            n_total = 0
            for f in uploaded:
                dest = CRM_DIR / f.name
                raw = f.read()
                # Compter le total avant import pour calculer les doublons ignorés
                try:
                    import io as _io
                    _df_count = pd.read_excel(_io.BytesIO(raw)) if f.name.endswith(".xlsx") else pd.read_csv(_io.BytesIO(raw))
                    n_total += len(_df_count)
                except Exception:
                    log.warning("Comptage lignes CRM echoue pour '%s'", f.name, exc_info=True)
                dest.write_bytes(raw)
                n_imp += _import_crm_file_to_db(dest)
            n_skip = n_total - n_imp
            msg = f"{len(uploaded)} fichier(s) ajouté(s) · {n_imp} entreprise(s) intégrée(s)"
            if n_skip > 0:
                msg += f" · {n_skip} doublon(s) ignoré(s)"
            st.success(msg + ".")
            st.rerun()

        st.divider()
        st.subheader("Fichiers actifs")
        if stats_crm["liste"]:
            for item in stats_crm["liste"]:
                c_n, c_nb, c_s, c_d = st.columns([4, 1, 1, 1])
                c_n.write(f"📄 {item['nom']}")
                c_nb.caption(f"{item['nb']} lignes")
                if c_s.button("Synchroniser", key=f"sync_{item['nom']}", type="secondary"):
                    n = _import_crm_file_to_db(Path(item["path"]))
                    st.success(f"{n} entreprise(s) synchronisée(s).")
                    st.rerun()
                if c_d.button("Supprimer", key=f"del_{item['nom']}", type="secondary"):
                    Path(item["path"]).unlink(missing_ok=True)
                    st.success(f"« {item['nom']} » supprimé.")
                    st.rerun()
        else:
            st.info("Aucun fichier CRM actif.")

        st.divider()
        st.subheader("Comparer un fichier")
        st.caption("Vérifie lesquelles de ces entreprises sont déjà dans ton CRM — sans l'ajouter à la base.")
        cmp_file = st.file_uploader(
            "Comparer", type=["xlsx", "csv"],
            key="crm_compare", label_visibility="collapsed",
        )
        if cmp_file:
            res = compare_against_crm(cmp_file, cmp_file.name)
            if res.get("error"):
                st.error(res["error"])
            else:
                r1, r2, r3 = st.columns(3)
                r1.metric("Total analysé",    res["total"])
                r2.metric("Déjà dans le CRM", len(res["doublons"]))
                r3.metric("Nouveaux",          len(res["nouveaux"]))
                if res["doublons"]:
                    st.markdown("**Doublons**")
                    st.dataframe(pd.DataFrame(res["doublons"]), use_container_width=True, hide_index=True)
                if res["nouveaux"]:
                    st.markdown("**Nouveaux**")
                    st.dataframe(pd.DataFrame(res["nouveaux"]), use_container_width=True, hide_index=True)
      except Exception as e:
        log.error("Erreur onglet CRM", exc_info=True)
        st.error(f"Erreur CRM : {e}")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — Tous les leads
# ══════════════════════════════════════════════════════════════════════════════
with tab5:
    try:
        st.title("Tous les leads")
        st.caption("Vue globale toutes sessions confondues.")

        df_all = load_df()
        if df_all.empty:
            st.info("Aucun lead en base.")
        else:
            nb_qual_all  = int(df_all.apply(is_qualifie, axis=1).sum())
            nb_email_all = int(df_all["email"].notna().sum()) if "email" in df_all.columns else 0
            a2_g = has_agent2(df_all)

            t1, t2, t3, t4, t5 = st.columns(5)
            t1.metric("Total",          len(df_all))
            t2.metric("Avec site",      int(df_all["website_url"].notna().sum()))
            t3.metric("Avec dirigeant", int(df_all["owner_name"].notna().sum()))
            t4.metric("Avec email",     nb_email_all)
            t5.metric("Qualifiés",      nb_qual_all)

            st.divider()

            fc1, fc2, fc3 = st.columns(3)
            with fc1:
                f_s = st.multiselect("Secteur", sorted(df_all["sector"].dropna().unique()), key="s4")
            with fc2:
                f_v = st.multiselect("Ville", sorted(df_all["city"].dropna().unique()), key="v4")
            with fc3:
                sess_all = LeadQueue(str(ROOT / config.db_path)).list_sessions()
                sid_map4 = {fmt_session(s): s["session_id"] for s in sess_all}
                f_se = st.multiselect("Session", list(sid_map4.keys()), key="se4")
                sids_sel = [sid_map4[l] for l in f_se]

            fc4, fc5, fc6 = st.columns(3)
            with fc4: f_q  = st.toggle("Qualifiés uniquement", key="fq4")
            with fc5: f_m  = st.checkbox("Avec email", key="fe4")
            with fc6: f_si = st.checkbox("Avec site",  key="fs4")

            df_f = df_all.copy()
            if f_s:      df_f = df_f[df_f["sector"].isin(f_s)]
            if f_v:      df_f = df_f[df_f["city"].isin(f_v)]
            if sids_sel: df_f = df_f[df_f["session_id"].isin(sids_sel)]
            if f_q:      df_f = df_f[df_f.apply(is_qualifie, axis=1)]
            if f_m:      df_f = df_f[df_f["email"].notna()]
            if f_si:     df_f = df_f[df_f["website_url"].notna()]

            st.caption(f"{len(df_f)} leads affichés")
            cols_v4 = [c for c in COLS_VIEW if c in df_f.columns]
            st.dataframe(df_f[cols_v4].rename(columns=COLS_VIEW),
                         use_container_width=True, hide_index=True, height=480)

            st.divider()
            ex1, ex2 = st.columns(2)
            with ex1:
                xl_g = to_excel_combined(df_f, COLS_A1, COLS_A2) if a2_g else to_excel_a1(df_f[[c for c in COLS_A1 if c in df_f.columns]].rename(columns=COLS_A1))
                st.download_button(f"Exporter la sélection ({len(df_f)})", data=xl_g,
                                   file_name="leads_global.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True, type="primary")
            with ex2:
                xl_gv = to_excel_vendeur(df_f)
                st.download_button(f"🎯  Vue vendeur ({len(df_f)})", data=xl_gv,
                                   file_name="leads_vendeur.xlsx",
                                   mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                   use_container_width=True)

            st.divider()
            with st.expander("Nettoyer les doublons"):
                st.caption("Analyse toute la base et supprime les entrées en double (même entreprise dans plusieurs sessions ou fichiers CRM).")
                dc1, dc2 = st.columns([3, 1])
                with dc1:
                    if st.button("Analyser les doublons", key="btn_count_dups", type="secondary", use_container_width=True):
                        with st.spinner("Analyse en cours…"):
                            nb_grp, nb_trop = _count_db_duplicates()
                        if nb_grp == 0:
                            st.success("Aucun doublon détecté.")
                        else:
                            st.warning(f"{nb_grp} groupe(s) de doublons · {nb_trop} entrée(s) en trop.")
                            st.session_state["_dup_found"] = (nb_grp, nb_trop)
                with dc2:
                    dup_info = st.session_state.get("_dup_found")
                    btn_label = f"Supprimer ({dup_info[1]})" if dup_info else "Supprimer"
                    if st.button(btn_label, key="btn_del_dups", type="primary",
                                 use_container_width=True, disabled=not dup_info):
                        with st.spinner("Suppression…"):
                            _, nb_del = _find_db_duplicates()
                        st.session_state.pop("_dup_found", None)
                        st.success(f"{nb_del} doublon(s) supprimé(s).")
                        st.rerun()
    except Exception as e:
        log.error("Erreur onglet Tous les leads", exc_info=True)
        st.error(f"Erreur : {e}")


@st.dialog("✉️ Nouveau message", width="large")
def _compose_email_dialog():
    _to = st.session_state.get("_compose_email_to", "")
    _lead_name = st.session_state.get("_compose_email_name", "")
    if not _to:
        st.warning("Aucun destinataire.")
        return
    if not is_gmail_configured():
        st.markdown(
            '<div style="background:#1A1700;border:1px solid #3A2E00;border-radius:8px;padding:12px 16px;font-size:13px;color:#F5D87A">'
            'Gmail non configuré — va dans l\'onglet <b>Configuration</b> pour ajouter ton adresse et mot de passe d\'application.</div>',
            unsafe_allow_html=True,
        )
        return
    _mail_cfg = load_user_config()
    _from_addr = _mail_cfg.get("gmail_address", "")

    # ── En-tête De / À ──
    st.markdown(
        f'<div style="padding:6px 0 10px;border-bottom:1px solid #2A2D35;margin-bottom:8px">'
        f'<div style="font-size:12px;color:#9CA3AF;margin-bottom:4px">De : <b style="color:#F0EBE3">{_from_addr}</b></div>'
        f'<div style="font-size:12px;color:#9CA3AF">À : <b style="color:#F0EBE3">{_to}</b></div></div>',
        unsafe_allow_html=True,
    )

    # ── CC / BCC (dépliable) ──
    _show_cc = st.checkbox("Ajouter Cc / Cci", key="_compose_show_cc")
    _cc = ""
    _bcc = ""
    if _show_cc:
        _cc_col, _bcc_col = st.columns(2)
        with _cc_col:
            _cc = st.text_input("Cc", key="_compose_cc", placeholder="email1@ex.com, email2@ex.com")
        with _bcc_col:
            _bcc = st.text_input("Cci (copie cachée)", key="_compose_bcc", placeholder="email@ex.com")

    # ── Objet ──
    _subj = st.text_input("Objet", value=f"Prise de contact — {_lead_name}", key="_compose_subject")

    # ── Corps du message ──
    _body = st.text_area("Message", height=200, key="_compose_body", placeholder="Bonjour,\n\nJe me permets de vous contacter...")

    # ── Pièces jointes ──
    _files = st.file_uploader(
        "Pièces jointes",
        accept_multiple_files=True,
        key="_compose_attachments",
        help="PDF, images, documents… — max 25 Mo au total (limite Gmail)",
    )
    if _files:
        _total_size = sum(f.size for f in _files)
        _size_mb = _total_size / (1024 * 1024)
        if _size_mb > 25:
            st.warning(f"Taille totale : {_size_mb:.1f} Mo — dépasse la limite Gmail de 25 Mo.")
        else:
            _names = ", ".join(f.name for f in _files)
            st.markdown(
                f'<div style="font-size:11px;color:#9CA3AF;padding:4px 0">📎 {len(_files)} fichier(s) — {_size_mb:.1f} Mo — {_names}</div>',
                unsafe_allow_html=True,
            )

    # ── Options ──
    _opt_col1, _opt_col2 = st.columns(2)
    with _opt_col1:
        _priority = st.selectbox(
            "Priorité",
            options=["normal", "high", "low"],
            format_func=lambda x: {"normal": "Normale", "high": "🔴 Haute", "low": "🔵 Basse"}[x],
            key="_compose_priority",
        )
    with _opt_col2:
        _send_html = st.checkbox("Envoyer en HTML", key="_compose_html", help="Interprète le message comme du HTML (gras, liens, listes…)")

    # ── Confirmation de lecture ──
    _read_receipt = st.checkbox("Demander un accusé de lecture", key="_compose_read_receipt")

    st.divider()

    # ── Boutons ──
    _col_cancel, _col_send = st.columns([1, 1])
    with _col_cancel:
        if st.button("Annuler", use_container_width=True):
            st.session_state.pop("_compose_email_to", None)
            st.rerun()
    with _col_send:
        if st.button("Envoyer", type="primary", use_container_width=True):
            if not _body.strip():
                st.warning("Le message ne peut pas être vide.")
            else:
                _attach_list = None
                if _files:
                    _attach_list = [{"name": f.name, "data": f.read()} for f in _files]
                with st.spinner("Envoi en cours…"):
                    _ok, _msg = send_email(
                        _to, _subj, _body,
                        cc=_cc, bcc=_bcc,
                        html=_send_html,
                        priority=_priority,
                        attachments=_attach_list,
                        read_receipt=_read_receipt,
                    )
                if _ok:
                    _hist_entry = f"{datetime.now().strftime('%d/%m/%Y %H:%M')} — Mail envoyé à {_to} : {_subj}"
                    _db_path = str(ROOT / config.db_path)
                    try:
                        _hc = sqlite3.connect(_db_path)
                        _matched = _hc.execute("SELECT id, lead_history, tags FROM leads WHERE email=? LIMIT 1", (_to,)).fetchone()
                        if _matched:
                            _old_hist = str(_matched[1] or "")
                            _old_tags = str(_matched[2] or "")
                            _new_hist = (_old_hist + "\n" + _hist_entry).strip()
                            _tag_set = [t.strip() for t in _old_tags.split("|") if t.strip()]
                            if "Mail envoyé" not in _tag_set:
                                _tag_set.append("Mail envoyé")
                            _hc.execute("UPDATE leads SET lead_history=?, tags=?, updated_at=? WHERE id=?",
                                        (_new_hist, "|".join(_tag_set), datetime.now().isoformat(), _matched[0]))
                            _hc.commit()
                        _hc.close()
                    except Exception:
                        pass
                    st.success(f"Email envoyé à {_to}")
                    import time as _t
                    _t.sleep(1.5)
                    st.session_state.pop("_compose_email_to", None)
                    st.rerun()
                else:
                    st.error(_msg)


@st.dialog("📞 Appel en cours", width="large")
def _call_dialog():
    _phone_to = st.session_state.get("_call_phone_to", "")
    _lead_name = st.session_state.get("_call_lead_name", "")
    _lead_id = st.session_state.get("_call_lead_id")
    if not _phone_to:
        st.warning("Aucun numéro.")
        return

    _cfg_p = load_user_config()
    _user_phone = _cfg_p.get("user_phone", "")

    # ── En-tête appel ──
    st.markdown(
        f'<div style="text-align:center;padding:16px 0 8px">'
        f'<div style="font-size:28px;font-weight:700;color:#F0EBE3;margin-bottom:4px">{_lead_name}</div>'
        f'<div style="font-size:20px;color:#5B9BD5;font-weight:600;letter-spacing:1px">{_phone_to}</div>'
        f'</div>',
        unsafe_allow_html=True,
    )
    if _user_phone:
        st.markdown(
            f'<div style="text-align:center;font-size:11px;color:#6B7280;margin-bottom:8px">Depuis : {_user_phone}</div>',
            unsafe_allow_html=True,
        )

    st.divider()

    # ── Options d'appel ──
    _twilio_ok = is_twilio_configured()
    _telnyx_ok = is_telnyx_configured()
    _tel_clean = _phone_to.replace(" ", "").replace(".", "")

    _call_tabs = ["📞 Appel gratuit (forfait)"]
    if _twilio_ok:
        _call_tabs.insert(0, "📱 Twilio")
    if _telnyx_ok:
        _call_tabs.insert(0, "📱 Telnyx")
    if not _twilio_ok and not _telnyx_ok:
        _call_tabs.insert(0, "📱 Appel in-app (VoIP)")

    _created_tabs = st.tabs(_call_tabs)
    _tab_idx = 0

    if _telnyx_ok:
        with _created_tabs[_tab_idx]:
            st.markdown(
                '<div style="text-align:center;font-size:12px;color:#2ECC71;margin:8px 0">✓ Telnyx configuré</div>',
                unsafe_allow_html=True,
            )
            if st.button("📱 Lancer l'appel via Telnyx", key="call_telnyx_btn", type="primary", use_container_width=True):
                with st.spinner("Connexion en cours…"):
                    _ok, _msg = make_call_telnyx(_phone_to)
                if _ok:
                    st.success(_msg)
                    st.session_state["_call_started"] = True
                    st.session_state["_call_start_time"] = datetime.now().isoformat()
                    st.rerun()
                else:
                    st.error(_msg)
            st.markdown(
                '<div style="text-align:center;font-size:11px;color:#6B7280;margin-top:6px">'
                'Telnyx appelle ton téléphone, puis te connecte au lead · ~0.006€/min</div>',
                unsafe_allow_html=True,
            )
        _tab_idx += 1

    if _twilio_ok:
        with _created_tabs[_tab_idx]:
            st.markdown(
                '<div style="text-align:center;font-size:12px;color:#2ECC71;margin:8px 0">✓ Twilio configuré</div>',
                unsafe_allow_html=True,
            )
            if st.button("📱 Lancer l'appel via Twilio", key="call_twilio_btn", type="primary", use_container_width=True):
                with st.spinner("Connexion en cours…"):
                    _ok, _msg = make_call_twilio(_phone_to)
                if _ok:
                    st.success(_msg)
                    st.session_state["_call_started"] = True
                    st.session_state["_call_start_time"] = datetime.now().isoformat()
                    st.rerun()
                else:
                    st.error(_msg)
            st.markdown(
                '<div style="text-align:center;font-size:11px;color:#6B7280;margin-top:6px">'
                'Twilio appelle ton téléphone, puis te connecte au lead · ~0.013€/min</div>',
                unsafe_allow_html=True,
            )
        _tab_idx += 1

    if not _twilio_ok and not _telnyx_ok:
        with _created_tabs[_tab_idx]:
            st.markdown(
                '<div style="background:#111218;border:1px solid #2A2D35;border-radius:10px;padding:20px;text-align:center;opacity:0.6">'
                '<div style="font-size:32px;margin-bottom:8px">📱</div>'
                '<div style="font-size:14px;font-weight:700;color:#6B7280;margin-bottom:4px;text-decoration:line-through">'
                'Appel in-app (VoIP)</div>'
                '<div style="font-size:11px;color:#4A4D58;margin-bottom:12px">'
                'Passe des appels directement depuis LeadsEngine.<br>'
                'Le provider appelle ton téléphone → te connecte au lead → tout est loggé automatiquement.</div>'
                '<div style="display:inline-block;background:#E87B2A22;border:1px solid #E87B2A44;border-radius:6px;padding:6px 14px">'
                '<span style="font-size:11px;color:#E87B2A;font-weight:600">Configure Twilio ou Telnyx dans l\'onglet Configuration</span></div>'
                '<div style="font-size:10px;color:#4A4D58;margin-top:8px">Twilio ~0.013€/min · Telnyx ~0.006€/min</div>'
                '</div>',
                unsafe_allow_html=True,
            )
        _tab_idx += 1

    with _created_tabs[_tab_idx]:
        st.markdown(
            f'<div style="text-align:center;margin:12px 0">'
            f'<a href="tel:{_tel_clean}" target="_blank" style="display:inline-block;background:#2ECC71;color:#fff;'
            f'font-size:16px;font-weight:700;padding:14px 40px;border-radius:50px;text-decoration:none;'
            f'box-shadow:0 4px 15px #2ECC7144;letter-spacing:0.5px">'
            f'📞 Appeler {_tel_clean}</a></div>'
            f'<div style="text-align:center;font-size:11px;color:#6B7280;margin-top:8px">'
            f'Ouvre Phone Link ou ton application téléphone — utilise ton forfait mobile</div>',
            unsafe_allow_html=True,
        )
        st.markdown(
            '<div style="background:#111218;border:1px solid #2A2D35;border-radius:8px;padding:10px 16px;margin-top:10px;font-size:11px;color:#6B7280">'
            '<b style="color:#9CA3AF">Astuce :</b> Installe '
            '<a href="ms-windows-store://pdp/?productid=9NMPJ99VJBWV" target="_blank" style="color:#5B9BD5">Phone Link</a>'
            ' (gratuit, Microsoft) pour appeler depuis ton PC via ton smartphone.</div>',
            unsafe_allow_html=True,
        )

    st.divider()

    # ── Chronomètre (manuel) ──
    if "_call_started" not in st.session_state:
        st.session_state["_call_started"] = False
        st.session_state["_call_start_time"] = None

    _timer_col1, _timer_col2 = st.columns([1, 1])
    with _timer_col1:
        if not st.session_state["_call_started"]:
            if st.button("Démarrer le chrono", key="call_start_timer", use_container_width=True):
                st.session_state["_call_started"] = True
                st.session_state["_call_start_time"] = datetime.now().isoformat()
                st.rerun()
        else:
            _start_dt = datetime.fromisoformat(st.session_state["_call_start_time"])
            _elapsed = datetime.now() - _start_dt
            _mins = int(_elapsed.total_seconds() // 60)
            _secs = int(_elapsed.total_seconds() % 60)
            st.markdown(
                f'<div style="text-align:center;padding:8px;background:#0F1016;border-radius:8px">'
                f'<div style="font-size:10px;color:#6B7280;margin-bottom:2px">DURÉE</div>'
                f'<div style="font-size:24px;font-weight:700;color:#F0EBE3;font-family:monospace">{_mins:02d}:{_secs:02d}</div></div>',
                unsafe_allow_html=True,
            )
    with _timer_col2:
        if st.session_state["_call_started"]:
            if st.button("Rafraîchir", key="call_refresh_timer", use_container_width=True):
                st.rerun()

    st.divider()

    # ── Notes d'appel ──
    st.markdown(
        '<div style="font-size:12px;font-weight:600;color:#E87B2A;margin-bottom:6px">NOTES D\'APPEL</div>',
        unsafe_allow_html=True,
    )
    _call_notes = st.text_area(
        "Notes",
        height=120,
        key="_call_notes",
        placeholder="Points abordés, suite à donner, impression…",
        label_visibility="collapsed",
    )

    # ── Résultat de l'appel ──
    _call_result = st.selectbox(
        "Résultat de l'appel",
        options=["a_rappeler", "en_cours", "interesse", "refus", "non_appele"],
        format_func=lambda x: {
            "non_appele": "Non abouti",
            "a_rappeler": "À rappeler",
            "en_cours": "En discussion",
            "interesse": "Intéressé",
            "refus": "Refus / Pas intéressé",
        }[x],
        key="_call_result",
    )

    st.divider()

    # ── Boutons ──
    _bc1, _bc2 = st.columns([1, 1])
    with _bc1:
        if st.button("Fermer sans sauvegarder", use_container_width=True, key="call_cancel"):
            st.session_state.pop("_call_phone_to", None)
            st.session_state.pop("_call_started", None)
            st.session_state.pop("_call_start_time", None)
            st.rerun()
    with _bc2:
        if st.button("Enregistrer et fermer", type="primary", use_container_width=True, key="call_save"):
            _duration_str = ""
            if st.session_state.get("_call_started") and st.session_state.get("_call_start_time"):
                _start_dt = datetime.fromisoformat(st.session_state["_call_start_time"])
                _el = datetime.now() - _start_dt
                _m = int(_el.total_seconds() // 60)
                _s = int(_el.total_seconds() % 60)
                _duration_str = f"{_m}min{_s:02d}s"

            _existing_notes = ""
            if _lead_id is not None:
                _db_path = str(ROOT / config.db_path)
                _cn = sqlite3.connect(_db_path)
                _row_n = _cn.execute("SELECT lead_notes FROM leads WHERE id=?", (int(_lead_id),)).fetchone()
                if _row_n and _row_n[0]:
                    _existing_notes = _row_n[0]
                _cn.close()

            _timestamp = datetime.now().strftime("%d/%m/%Y %H:%M")
            _new_entry = f"[APPEL {_timestamp}]"
            if _duration_str:
                _new_entry += f" Durée: {_duration_str}"
            _result_labels = {
                "non_appele": "Non abouti", "a_rappeler": "À rappeler",
                "en_cours": "En discussion", "interesse": "Intéressé", "refus": "Refus",
            }
            _new_entry += f" | {_result_labels.get(_call_result, _call_result)}"
            if _call_notes.strip():
                _new_entry += f"\n  → {_call_notes.strip()}"

            _final_notes = f"{_new_entry}\n{_existing_notes}".strip() if _existing_notes else _new_entry

            if _lead_id is not None:
                _db_path = str(ROOT / config.db_path)
                _cn = sqlite3.connect(_db_path)
                _cn.execute(
                    "UPDATE leads SET call_status=?, lead_notes=?, updated_at=? WHERE id=?",
                    (_call_result, _final_notes, datetime.now().isoformat(), int(_lead_id)),
                )
                _cn.commit()
                _cn.close()

            st.success("Appel enregistré.")
            import time as _t2
            _t2.sleep(1)
            st.session_state.pop("_call_phone_to", None)
            st.session_state.pop("_call_started", None)
            st.session_state.pop("_call_start_time", None)
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# TAB 7 — Fiches Leads
# ══════════════════════════════════════════════════════════════════════════════
with tab7:
    if not _IS_PRO:
        st.markdown(
            '<div style="text-align:center;padding:80px 20px">'
            '<div style="font-size:48px;margin-bottom:16px">🔒</div>'
            '<div style="font-size:18px;font-weight:700;color:#F0EBE3;margin-bottom:8px">Fonctionnalité Pro</div>'
            '<div style="font-size:13px;color:#4A4D58">Les fiches leads avec scoring, email et appels sont réservées à la licence Pro.</div>'
            '<a href="https://leadsengine.netlify.app" target="_blank" '
            'style="display:inline-block;margin-top:20px;background:#E87B2A;color:#0D0E11;padding:8px 20px;'
            'border-radius:6px;font-weight:700;font-size:13px;text-decoration:none">Passer en Pro</a>'
            '</div>',
            unsafe_allow_html=True,
        )
    if st.session_state.get("_compose_email_to") and _IS_PRO:
        _compose_email_dialog()
    if st.session_state.get("_call_phone_to") and _IS_PRO:
        _call_dialog()

    if _IS_PRO:
      try:
        _title_c1, _title_c2 = st.columns([3, 1])
        with _title_c1:
            st.title("Fiches Leads")
        with _title_c2:
            st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
            if st.button("+ Ajouter un lead", key="btn_add_lead", type="secondary", use_container_width=True):
                st.session_state["_show_add_lead"] = True

        # ── Formulaire ajout manuel ──────────────────────────────
        if st.session_state.get("_show_add_lead"):
            with st.expander("Nouveau lead", expanded=True):
                _al_c1, _al_c2 = st.columns(2)
                with _al_c1:
                    _al_name = st.text_input("Nom entreprise *", key="al_name")
                    _al_city = st.text_input("Ville *", key="al_city")
                    _al_sector = st.text_input("Secteur", key="al_sector")
                    _al_phone = st.text_input("Téléphone", key="al_phone")
                with _al_c2:
                    _al_email = st.text_input("Email", key="al_email")
                    _al_web = st.text_input("Site web", key="al_web")
                    _al_owner = st.text_input("Dirigeant", key="al_owner")
                    _al_notes = st.text_input("Notes", key="al_notes")
                _al_btn1, _al_btn2 = st.columns(2)
                with _al_btn1:
                    if st.button("Enregistrer", type="primary", key="al_save", use_container_width=True):
                        if _al_name.strip() and _al_city.strip():
                            _db_path = str(ROOT / config.db_path)
                            _alc = sqlite3.connect(_db_path)
                            _alc.execute(
                                "INSERT INTO leads (company_name, city, sector, source, phone, email, website_url, owner_name, lead_notes, status, scraped_at, updated_at) "
                                "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                                (_al_name.strip(), _al_city.strip(), _al_sector.strip(), "manuel",
                                 _al_phone.strip() or None, _al_email.strip() or None,
                                 _al_web.strip() or None, _al_owner.strip() or None,
                                 _al_notes.strip() or None, "scraped",
                                 datetime.now().isoformat(), datetime.now().isoformat()),
                            )
                            _alc.commit(); _alc.close()
                            st.session_state.pop("_show_add_lead", None)
                            st.success("Lead ajouté.")
                            st.rerun()
                        else:
                            st.error("Nom et ville sont obligatoires.")
                with _al_btn2:
                    if st.button("Annuler", key="al_cancel", use_container_width=True):
                        st.session_state.pop("_show_add_lead", None)
                        st.rerun()

        df_fiches = load_df()
        if df_fiches.empty:
            st.info("Aucun lead en base.")
        else:
            df_fiches["_score"] = df_fiches.apply(compute_lead_score, axis=1)
            df_fiches = df_fiches.sort_values("_score", ascending=False).reset_index(drop=True)

            # ── Filtres ──────────────────────────────────────────────────────
            ff1, ff2, ff3, ff4 = st.columns(4)
            with ff1:
                _f_search = st.text_input("Rechercher", placeholder="Nom, ville, secteur…", key="fiche_search")
            with ff2:
                _f_score = st.select_slider("Score min.", options=list(range(0, 101, 10)), value=0, key="fiche_score_min")
            with ff3:
                _f_sector = st.multiselect("Secteur", sorted(df_fiches["sector"].dropna().unique()), key="fiche_sector")
            with ff4:
                _f_city = st.multiselect("Ville", sorted(df_fiches["city"].dropna().unique()), key="fiche_city")

            df_ff = df_fiches.copy()
            if _f_search:
                _q = _f_search.lower()
                df_ff = df_ff[
                    df_ff["company_name"].fillna("").str.lower().str.contains(_q, na=False) |
                    df_ff["city"].fillna("").str.lower().str.contains(_q, na=False) |
                    df_ff["sector"].fillna("").str.lower().str.contains(_q, na=False) |
                    df_ff["owner_name"].fillna("").str.lower().str.contains(_q, na=False)
                ]
            if _f_score > 0:
                df_ff = df_ff[df_ff["_score"] >= _f_score]
            if _f_sector:
                df_ff = df_ff[df_ff["sector"].isin(_f_sector)]
            if _f_city:
                df_ff = df_ff[df_ff["city"].isin(_f_city)]

            # ── Stats rapides (cliquables = filtres) ────────────────────────
            _nb_green = int((df_ff["_score"] >= 60).sum())
            _nb_yellow = int(((df_ff["_score"] >= 30) & (df_ff["_score"] < 60)).sum())
            _nb_red = int((df_ff["_score"] < 30).sum())

            if "_fiche_filter" not in st.session_state:
                st.session_state["_fiche_filter"] = "all"
            _active_filter = st.session_state["_fiche_filter"]

            _s1, _s2, _s3, _s4 = st.columns(4)
            with _s1:
                _sel_all = "border:2px solid #E87B2A" if _active_filter == "all" else "border:1px solid #1E2028"
                st.markdown(f'<div style="background:#13141A;{_sel_all};border-radius:10px;padding:14px 18px;text-align:center">'
                    f'<div style="font-size:24px;font-weight:800;color:#F0EBE3">{len(df_ff)}</div>'
                    f'<div style="font-size:11px;color:#6B7280;margin-top:2px">Total</div></div>', unsafe_allow_html=True)
                if st.button("Tous", key="fiche_filter_all", use_container_width=True, type="tertiary"):
                    st.session_state["_fiche_filter"] = "all"
                    st.session_state["_fiche_page"] = 0
                    st.session_state.pop("_fiche_selected_id", None)
                    st.rerun()
            with _s2:
                _sel_bon = "border:2px solid #2ECC71" if _active_filter == "bon" else "border:1px solid #1A3320"
                st.markdown(f'<div style="background:#0D1A0F;{_sel_bon};border-radius:10px;padding:14px 18px;text-align:center">'
                    f'<div style="font-size:24px;font-weight:800;color:#2ECC71">{_nb_green}</div>'
                    f'<div style="font-size:11px;color:#2ECC71;margin-top:2px">Bon ≥60</div></div>', unsafe_allow_html=True)
                if st.button("Bon", key="fiche_filter_bon", use_container_width=True, type="tertiary"):
                    st.session_state["_fiche_filter"] = "bon"
                    st.session_state["_fiche_page"] = 0
                    st.session_state.pop("_fiche_selected_id", None)
                    st.rerun()
            with _s3:
                _sel_moy = "border:2px solid #F5D87A" if _active_filter == "moyen" else "border:1px solid #3A2E00"
                st.markdown(f'<div style="background:#1A1700;{_sel_moy};border-radius:10px;padding:14px 18px;text-align:center">'
                    f'<div style="font-size:24px;font-weight:800;color:#F5D87A">{_nb_yellow}</div>'
                    f'<div style="font-size:11px;color:#F5D87A;margin-top:2px">Moyen 30-59</div></div>', unsafe_allow_html=True)
                if st.button("Moyen", key="fiche_filter_moyen", use_container_width=True, type="tertiary"):
                    st.session_state["_fiche_filter"] = "moyen"
                    st.session_state["_fiche_page"] = 0
                    st.session_state.pop("_fiche_selected_id", None)
                    st.rerun()
            with _s4:
                _sel_mau = "border:2px solid #E74C3C" if _active_filter == "mauvais" else "border:1px solid #3A1A1A"
                st.markdown(f'<div style="background:#1A0D0D;{_sel_mau};border-radius:10px;padding:14px 18px;text-align:center">'
                    f'<div style="font-size:24px;font-weight:800;color:#E74C3C">{_nb_red}</div>'
                    f'<div style="font-size:11px;color:#E74C3C;margin-top:2px">Mauvais &lt;30</div></div>', unsafe_allow_html=True)
                if st.button("Mauvais", key="fiche_filter_mauvais", use_container_width=True, type="tertiary"):
                    st.session_state["_fiche_filter"] = "mauvais"
                    st.session_state["_fiche_page"] = 0
                    st.session_state.pop("_fiche_selected_id", None)
                    st.rerun()

            # Applique le filtre de catégorie
            if _active_filter == "bon":
                df_ff = df_ff[df_ff["_score"] >= 60]
            elif _active_filter == "moyen":
                df_ff = df_ff[(df_ff["_score"] >= 30) & (df_ff["_score"] < 60)]
            elif _active_filter == "mauvais":
                df_ff = df_ff[df_ff["_score"] < 30]

            # Statuts d'appel
            CALL_STATUSES = {
                "non_appele": ("Non appelé", "#6B7280"),
                "a_rappeler": ("À rappeler", "#F5D87A"),
                "en_cours": ("En cours", "#5B9BD5"),
                "interesse": ("Intéressé", "#2ECC71"),
                "refus": ("Refus", "#E74C3C"),
            }

            # Filtre par statut d'appel
            _fc_status_col1, _fc_status_col2 = st.columns([1, 3])
            with _fc_status_col1:
                _call_filter_options = ["Tous"] + [v[0] for v in CALL_STATUSES.values()]
                _call_filter = st.selectbox("Statut appel", _call_filter_options, key="fiche_call_filter")
            if _call_filter != "Tous":
                _call_key = [k for k, v in CALL_STATUSES.items() if v[0] == _call_filter][0]
                df_ff = df_ff[df_ff["call_status"].fillna("non_appele") == _call_key]

            # ── Grille de cartes (liste côté gauche) + fiche détaillée (droite) ─
            _card_col, _detail_col = st.columns([2, 3])

            with _card_col:
                st.markdown(f'<div style="font-size:12px;color:#6B7280;margin-bottom:8px">{len(df_ff)} leads</div>', unsafe_allow_html=True)

                # Pagination
                _per_page = 15
                _total_pages = max(1, (len(df_ff) + _per_page - 1) // _per_page)
                if "_fiche_page" not in st.session_state:
                    st.session_state["_fiche_page"] = 0
                _page = st.session_state["_fiche_page"]
                _start = _page * _per_page
                _end = min(_start + _per_page, len(df_ff))
                df_page = df_ff.iloc[_start:_end]

                for _idx, _row in df_page.iterrows():
                    _sc = _row["_score"]
                    _emoji, _color = score_label(_sc)
                    _name = str(_row.get("company_name") or "—")
                    _city_v = str(_row.get("city") or "")
                    _sector_v = str(_row.get("sector") or "")
                    _dirigeant = str(_row.get("owner_name") or "")
                    _email_v = str(_row.get("email")) if pd.notna(_row.get("email")) else ""
                    _phone_v = str(_row.get("phone")) if pd.notna(_row.get("phone")) else ""
                    _web_v = str(_row.get("website_url")) if pd.notna(_row.get("website_url")) else ""
                    _grating = _row.get("google_rating")
                    _grating_str = f"{float(_grating):.1f}★" if pd.notna(_grating) else ""
                    _lead_db_id = _row.get("id")

                    # Statut appel
                    _call_st = str(_row.get("call_status") or "non_appele")
                    if _call_st not in CALL_STATUSES:
                        _call_st = "non_appele"
                    _call_label, _call_color = CALL_STATUSES[_call_st]

                    # Notes aperçu
                    _notes_raw = str(_row.get("lead_notes") or "") if pd.notna(_row.get("lead_notes")) else ""
                    _notes_preview = _notes_raw.replace("\n", " ").strip()
                    if len(_notes_preview) > 60:
                        _notes_preview = _notes_preview[:57] + "…"

                    # Tags ligne 1 : statut + contact
                    _tags_html = f'<span style="background:{_call_color}22;color:{_call_color};font-size:9px;font-weight:700;padding:2px 6px;border-radius:3px;margin-right:4px;border:1px solid {_call_color}44">{_call_label.upper()}</span>'
                    if _email_v:
                        _tags_html += '<span style="background:#1A2332;color:#5B9BD5;font-size:9px;padding:2px 6px;border-radius:3px;margin-right:4px">EMAIL</span>'
                    if _phone_v:
                        _tags_html += '<span style="background:#1A2332;color:#5B9BD5;font-size:9px;padding:2px 6px;border-radius:3px;margin-right:4px">TEL</span>'
                    if _web_v:
                        _tags_html += '<span style="background:#1A2332;color:#5B9BD5;font-size:9px;padding:2px 6px;border-radius:3px;margin-right:4px">SITE</span>'
                    if _grating_str:
                        _tags_html += f'<span style="background:#1A2010;color:#A8D86E;font-size:9px;padding:2px 6px;border-radius:3px">{_grating_str}</span>'

                    # Notes aperçu ligne
                    _notes_html = ""
                    if _notes_preview:
                        _notes_html = f'<div style="font-size:10px;color:#8B8070;margin-top:4px;font-style:italic;overflow:hidden;text-overflow:ellipsis;white-space:nowrap">📝 {_notes_preview}</div>'

                    _selected = st.session_state.get("_fiche_selected_id") == _lead_db_id
                    _border_color = _color if _selected else "#1E2028"
                    _bg = "#161820" if _selected else "#0F1016"

                    # Carte complète cliquable via bouton
                    if st.button(
                        f"{_emoji} {_sc}  ·  {_name}  —  {_city_v}" if _city_v else f"{_emoji} {_sc}  ·  {_name}",
                        key=f"fiche_btn_{_idx}",
                        use_container_width=True,
                        type="primary" if _selected else "tertiary",
                    ):
                        st.session_state["_fiche_selected_id"] = _lead_db_id
                        st.rerun()
                    # Détails visuels sous le bouton
                    st.markdown(
                        f'<div style="background:{_bg};border:1px solid {_border_color};border-left:3px solid {_color};'
                        f'border-radius:0 0 8px 8px;padding:6px 14px 8px;margin-top:-16px;margin-bottom:8px">'
                        f'<div style="font-size:11px;color:#6B7280">'
                        f'{_sector_v}{" · " + _dirigeant if _dirigeant else ""}</div>'
                        f'<div style="margin-top:4px">{_tags_html}</div>'
                        f'{_notes_html}'
                        f'</div>',
                        unsafe_allow_html=True,
                    )

                # Pagination controls
                if _total_pages > 1:
                    _p1, _p2, _p3 = st.columns([1, 2, 1])
                    with _p1:
                        if st.button("◀", key="fiche_prev", disabled=_page == 0, use_container_width=True):
                            st.session_state["_fiche_page"] = _page - 1
                            st.session_state.pop("_fiche_selected_id", None)
                            st.rerun()
                    with _p2:
                        st.markdown(f'<div style="text-align:center;color:#6B7280;font-size:12px;padding:8px 0">Page {_page + 1} / {_total_pages}</div>', unsafe_allow_html=True)
                    with _p3:
                        if st.button("▶", key="fiche_next", disabled=_page >= _total_pages - 1, use_container_width=True):
                            st.session_state["_fiche_page"] = _page + 1
                            st.session_state.pop("_fiche_selected_id", None)
                            st.rerun()

            # ── Fiche détaillée (panneau droit) ──────────────────────────────
            with _detail_col:
                _sel_db_id = st.session_state.get("_fiche_selected_id")
                _r = None
                if _sel_db_id is not None and "id" in df_ff.columns:
                    _match = df_ff[df_ff["id"] == _sel_db_id]
                    if not _match.empty:
                        _r = _match.iloc[0]
                if _r is not None:
                    _sc = _r["_score"]
                    _emoji, _color = score_label(_sc)
                    _name = str(_r.get("company_name") or "—")

                    # Header fiche
                    st.markdown(
                        f'<div style="background:#13141A;border:1px solid #1E2028;border-radius:12px;padding:20px 24px;margin-bottom:16px">'
                        f'<div style="display:flex;justify-content:space-between;align-items:flex-start">'
                        f'<div>'
                        f'<div style="font-size:20px;font-weight:800;color:#F0EBE3">{_name}</div>'
                        f'<div style="font-size:12px;color:#6B7280;margin-top:4px">'
                        f'{_r.get("city") or ""}{" · " + str(_r.get("sector") or "") if _r.get("sector") else ""}</div>'
                        f'</div>'
                        f'<div style="text-align:center">'
                        f'<div style="background:{_color}18;border:2px solid {_color};border-radius:50%;width:56px;height:56px;'
                        f'display:flex;align-items:center;justify-content:center;margin:0 auto">'
                        f'<span style="font-size:20px;font-weight:900;color:{_color}">{_sc}</span></div>'
                        f'<div style="font-size:10px;color:{_color};margin-top:4px">SCORE</div>'
                        f'</div></div>'
                        f'<div style="margin-top:12px;background:#0F1016;border-radius:6px;height:8px;overflow:hidden">'
                        f'<div style="width:{_sc}%;height:100%;background:linear-gradient(90deg,{_color}88,{_color});border-radius:6px"></div>'
                        f'</div></div>',
                        unsafe_allow_html=True,
                    )

                    # ── Bouton Modifier ──────────────────────────────────────
                    if st.button("✏️ Modifier ce lead", key=f"edit_lead_{_sel_db_id}", use_container_width=True):
                        st.session_state["_edit_lead_id"] = _sel_db_id

                    if st.session_state.get("_edit_lead_id") == _sel_db_id:
                        with st.expander("Modifier le lead", expanded=True):
                            _ed_c1, _ed_c2 = st.columns(2)
                            with _ed_c1:
                                _ed_name = st.text_input("Nom", value=str(_r.get("company_name") or ""), key=f"ed_name_{_sel_db_id}")
                                _ed_city = st.text_input("Ville", value=str(_r.get("city") or ""), key=f"ed_city_{_sel_db_id}")
                                _ed_sector = st.text_input("Secteur", value=str(_r.get("sector") or ""), key=f"ed_sector_{_sel_db_id}")
                                _ed_phone = st.text_input("Téléphone", value=str(_r.get("phone") or "") if pd.notna(_r.get("phone")) else "", key=f"ed_phone_{_sel_db_id}")
                            with _ed_c2:
                                _ed_email = st.text_input("Email", value=str(_r.get("email") or "") if pd.notna(_r.get("email")) else "", key=f"ed_email_{_sel_db_id}")
                                _ed_web = st.text_input("Site web", value=str(_r.get("website_url") or "") if pd.notna(_r.get("website_url")) else "", key=f"ed_web_{_sel_db_id}")
                                _ed_owner = st.text_input("Dirigeant", value=str(_r.get("owner_name") or "") if pd.notna(_r.get("owner_name")) else "", key=f"ed_owner_{_sel_db_id}")
                                _ed_addr = st.text_input("Adresse", value=str(_r.get("address") or "") if pd.notna(_r.get("address")) else "", key=f"ed_addr_{_sel_db_id}")
                            _ed_b1, _ed_b2 = st.columns(2)
                            with _ed_b1:
                                if st.button("Sauvegarder", type="primary", key=f"ed_save_{_sel_db_id}", use_container_width=True):
                                    _db_path = str(ROOT / config.db_path)
                                    _edc = sqlite3.connect(_db_path)
                                    _edc.execute(
                                        "UPDATE leads SET company_name=?, city=?, sector=?, phone=?, email=?, website_url=?, owner_name=?, address=?, updated_at=? WHERE id=?",
                                        (_ed_name.strip(), _ed_city.strip(), _ed_sector.strip(),
                                         _ed_phone.strip() or None, _ed_email.strip() or None,
                                         _ed_web.strip() or None, _ed_owner.strip() or None,
                                         _ed_addr.strip() or None,
                                         datetime.now().isoformat(), int(_sel_db_id)),
                                    )
                                    _edc.commit(); _edc.close()
                                    st.session_state.pop("_edit_lead_id", None)
                                    st.success("Lead modifié.")
                                    st.rerun()
                            with _ed_b2:
                                if st.button("Annuler", key=f"ed_cancel_{_sel_db_id}", use_container_width=True):
                                    st.session_state.pop("_edit_lead_id", None)
                                    st.rerun()

                    # ── Section Contact ──────────────────────────────────────
                    st.markdown(
                        '<div style="font-size:13px;font-weight:700;color:#E87B2A;margin:16px 0 8px;letter-spacing:1px">CONTACT</div>',
                        unsafe_allow_html=True,
                    )
                    _contact_items = []
                    _dirigeant = str(_r.get("owner_name") or "")
                    _role = str(_r.get("owner_role") or "") if pd.notna(_r.get("owner_role")) else ""
                    _email_d = str(_r.get("email")) if pd.notna(_r.get("email")) else ""
                    _phone_d = str(_r.get("phone")) if pd.notna(_r.get("phone")) else ""
                    _web_d = str(_r.get("website_url")) if pd.notna(_r.get("website_url")) else ""

                    def _detail_row(icon, label, value, sub=""):
                        _s = f'<div style="display:flex;align-items:center;gap:10px;padding:8px 12px;background:#0F1016;border-radius:6px;margin-bottom:4px">'
                        _s += f'<span style="font-size:16px">{icon}</span>'
                        _s += f'<div><div style="font-size:12px;font-weight:600;color:#F0EBE3">{value}</div>'
                        _s += f'<div style="font-size:10px;color:#6B7280">{label}{" · " + sub if sub else ""}</div></div></div>'
                        return _s

                    _c_html = ""
                    if _dirigeant:
                        _c_html += _detail_row("👤", "Dirigeant", _dirigeant, _role)
                    if _email_d:
                        _c_html += _detail_row("✉️", "Email", _email_d)
                    if _phone_d:
                        _c_html += _detail_row("📞", "Téléphone", _phone_d)
                    if _web_d:
                        _href = _web_d if _web_d.startswith("http") else f"https://{_web_d}"
                        _c_html += _detail_row("🌐", "Site web", f'<a href="{_href}" target="_blank" style="color:#5B9BD5;text-decoration:none">{_web_d}</a>')
                    if not _c_html:
                        _c_html = '<div style="color:#6B7280;font-size:12px;padding:8px">Aucune donnée contact</div>'
                    st.markdown(_c_html, unsafe_allow_html=True)
                    _action_cols = []
                    if _email_d and _phone_d:
                        _action_cols = st.columns(2)
                    elif _email_d or _phone_d:
                        _action_cols = [st.columns(1)[0]]

                    _col_idx = 0
                    if _email_d:
                        _ctx = _action_cols[_col_idx] if len(_action_cols) > 1 else (_action_cols[0] if _action_cols else st)
                        with _ctx:
                            if st.button(f"✉️ Email", key=f"email_open_{_sel_db_id}", use_container_width=True):
                                st.session_state["_compose_email_to"] = _email_d
                                st.session_state["_compose_email_name"] = _name
                                st.rerun()
                        _col_idx += 1
                    if _phone_d:
                        _ctx = _action_cols[_col_idx] if len(_action_cols) > 1 else (_action_cols[0] if _action_cols else st)
                        with _ctx:
                            if st.button(f"📞 Appeler", key=f"call_open_{_sel_db_id}", use_container_width=True):
                                st.session_state["_call_phone_to"] = _phone_d
                                st.session_state["_call_lead_name"] = _name
                                st.session_state["_call_lead_id"] = _sel_db_id
                                st.rerun()

                    # ── Section Entreprise ───────────────────────────────────
                    st.markdown(
                        '<div style="font-size:13px;font-weight:700;color:#E87B2A;margin:20px 0 8px;letter-spacing:1px">ENTREPRISE</div>',
                        unsafe_allow_html=True,
                    )
                    _e_html = ""
                    _siren_d = str(_r.get("siren") or "") if pd.notna(_r.get("siren")) else ""
                    _siret_d = str(_r.get("siret") or "") if pd.notna(_r.get("siret")) else ""
                    _form_d = str(_r.get("legal_form") or "") if pd.notna(_r.get("legal_form")) else ""
                    _eff_d = str(_r.get("employee_range") or "") if pd.notna(_r.get("employee_range")) else ""
                    _crea_d = str(_r.get("creation_date") or "") if pd.notna(_r.get("creation_date")) else ""
                    _naf_d = str(_r.get("naf_label") or "") if pd.notna(_r.get("naf_label")) else ""
                    _addr_d = str(_r.get("address") or _r.get("siege_adresse") or "")
                    if not pd.notna(_addr_d) or not _addr_d.strip():
                        _addr_d = ""

                    if _siren_d:
                        _e_html += _detail_row("🏛️", "SIREN", _siren_d, f"SIRET {_siret_d}" if _siret_d else "")
                    if _form_d:
                        _e_html += _detail_row("📄", "Forme juridique", _form_d)
                    if _eff_d:
                        _e_html += _detail_row("👥", "Effectif", _eff_d)
                    if _crea_d:
                        _e_html += _detail_row("📅", "Date de création", _crea_d)
                    if _naf_d:
                        _e_html += _detail_row("🏷️", "Activité NAF", _naf_d)
                    if _addr_d:
                        _e_html += _detail_row("📍", "Adresse", _addr_d, str(_r.get("city") or ""))
                    if not _e_html:
                        _e_html = '<div style="color:#6B7280;font-size:12px;padding:8px">Aucune donnée entreprise</div>'
                    st.markdown(_e_html, unsafe_allow_html=True)

                    # ── Section Digital ───────────────────────────────────────
                    st.markdown(
                        '<div style="font-size:13px;font-weight:700;color:#E87B2A;margin:20px 0 8px;letter-spacing:1px">DIGITAL</div>',
                        unsafe_allow_html=True,
                    )
                    _d_html = ""
                    _grating_d = _r.get("google_rating")
                    if pd.notna(_grating_d):
                        _rev_d = int(_r.get("review_count")) if pd.notna(_r.get("review_count")) else 0
                        _stars = "★" * int(float(_grating_d)) + "☆" * (5 - int(float(_grating_d)))
                        _d_html += _detail_row("⭐", "Google Maps", f"{float(_grating_d):.1f} {_stars}", f"{_rev_d} avis")
                    _cms_d = str(_r.get("cms") or "") if pd.notna(_r.get("cms")) else ""
                    _host_d = str(_r.get("hosting") or "") if pd.notna(_r.get("hosting")) else ""
                    if _cms_d:
                        _d_html += _detail_row("🔧", "CMS", _cms_d, _host_d if _host_d else "")
                    elif _host_d:
                        _d_html += _detail_row("🖥️", "Hébergeur", _host_d)
                    _speed_d = _r.get("pagespeed_mobile")
                    if pd.notna(_speed_d):
                        _sp_val = int(_speed_d)
                        _sp_c = "#2ECC71" if _sp_val >= 70 else "#F5D87A" if _sp_val >= 50 else "#E74C3C"
                        _d_html += f'<div style="display:flex;align-items:center;gap:10px;padding:8px 12px;background:#0F1016;border-radius:6px;margin-bottom:4px">'
                        _d_html += f'<span style="font-size:16px">🚀</span>'
                        _d_html += f'<div style="flex:1"><div style="font-size:12px;font-weight:600;color:#F0EBE3">Vitesse mobile : {_sp_val}/100</div>'
                        _d_html += f'<div style="background:#1E2028;border-radius:4px;height:6px;margin-top:4px;overflow:hidden">'
                        _d_html += f'<div style="width:{_sp_val}%;height:100%;background:{_sp_c};border-radius:4px"></div></div></div></div>'
                    _seo_d = _r.get("seo_score")
                    if pd.notna(_seo_d):
                        _seo_val = int(_seo_d)
                        _seo_c = "#2ECC71" if _seo_val >= 7 else "#F5D87A" if _seo_val >= 4 else "#E74C3C"
                        _d_html += f'<div style="display:flex;align-items:center;gap:10px;padding:8px 12px;background:#0F1016;border-radius:6px;margin-bottom:4px">'
                        _d_html += f'<span style="font-size:16px">📊</span>'
                        _d_html += f'<div style="flex:1"><div style="font-size:12px;font-weight:600;color:#F0EBE3">Score SEO : {_seo_val}/10</div>'
                        _d_html += f'<div style="background:#1E2028;border-radius:4px;height:6px;margin-top:4px;overflow:hidden">'
                        _d_html += f'<div style="width:{_seo_val * 10}%;height:100%;background:{_seo_c};border-radius:4px"></div></div></div></div>'
                    _weak_d = str(_r.get("seo_weaknesses") or "") if pd.notna(_r.get("seo_weaknesses")) else ""
                    if _weak_d:
                        _d_html += _detail_row("⚠️", "Faiblesses SEO", _weak_d)
                    _age_d = str(_r.get("domain_age") or "") if pd.notna(_r.get("domain_age")) else ""
                    if _age_d:
                        _d_html += _detail_row("🕐", "Âge du domaine", _age_d)
                    _ads_d = _r.get("has_google_ads")
                    if pd.notna(_ads_d):
                        _ads_val = "Oui" if str(_ads_d) in ("1", "True", "true") else "Non"
                        _ads_icon = "💰" if _ads_val == "Oui" else "🚫"
                        _d_html += _detail_row(_ads_icon, "Google Ads", _ads_val)
                    if not _d_html:
                        _d_html = '<div style="color:#6B7280;font-size:12px;padding:8px">Aucune donnée digitale — lance l\'analyse sites</div>'
                    st.markdown(_d_html, unsafe_allow_html=True)

                    # ── Section Suivi ────────────────────────────────────────
                    st.markdown(
                        '<div style="font-size:13px;font-weight:700;color:#E87B2A;margin:20px 0 8px;letter-spacing:1px">SUIVI</div>',
                        unsafe_allow_html=True,
                    )

                    _lead_id = _r.get("id")
                    _current_call_st = str(_r.get("call_status") or "non_appele")
                    if _current_call_st not in CALL_STATUSES:
                        _current_call_st = "non_appele"
                    _current_notes = str(_r.get("lead_notes") or "") if pd.notna(_r.get("lead_notes")) else ""
                    _current_tags = str(_r.get("tags") or "") if pd.notna(_r.get("tags")) else ""
                    _current_history = str(_r.get("lead_history") or "") if pd.notna(_r.get("lead_history")) else ""

                    _st_options = list(CALL_STATUSES.keys())
                    _st_labels = [v[0] for v in CALL_STATUSES.values()]
                    _st_idx = _st_options.index(_current_call_st) if _current_call_st in _st_options else 0

                    _suivi_c1, _suivi_c2 = st.columns([1, 1])
                    with _suivi_c1:
                        _new_status = st.selectbox(
                            "Statut",
                            options=_st_options,
                            format_func=lambda x: CALL_STATUSES[x][0],
                            index=_st_idx,
                            key=f"fiche_call_st_{_sel_db_id}",
                        )
                    with _suivi_c2:
                        _st_color = CALL_STATUSES.get(_new_status, ("", "#6B7280"))[1]
                        _st_label = CALL_STATUSES.get(_new_status, ("", ""))[0]
                        st.markdown(
                            f'<div style="margin-top:28px;background:{_st_color}22;color:{_st_color};'
                            f'font-size:12px;font-weight:700;padding:8px 14px;border-radius:6px;text-align:center;'
                            f'border:1px solid {_st_color}44">{_st_label}</div>',
                            unsafe_allow_html=True,
                        )

                    # ── Étiquettes ──────────────────────────────────────
                    _TAG_PRESETS = {
                        "Mail envoyé": "#4A90D9",
                        "Relance 1": "#5B9BD5",
                        "Relance 2": "#3A7BBF",
                        "Relance 3": "#2A5B9F",
                        "Intéressé": "#2ECC71",
                        "Pas intéressé": "#E74C3C",
                        "RDV pris": "#9B59B6",
                        "Devis envoyé": "#F39C12",
                        "Signé": "#27AE60",
                        "Perdu": "#95A5A6",
                        "À rappeler": "#E87B2A",
                        "Réponse reçue": "#1ABC9C",
                    }
                    _tag_list = [t.strip() for t in _current_tags.split("|") if t.strip()] if _current_tags else []

                    _tags_display = ""
                    for _tg in _tag_list:
                        _tg_color = _TAG_PRESETS.get(_tg, "#6B7280")
                        _tags_display += (
                            f'<span style="background:{_tg_color}22;color:{_tg_color};font-size:11px;font-weight:700;'
                            f'padding:3px 10px;border-radius:4px;margin-right:4px;border:1px solid {_tg_color}44">{_tg}</span>'
                        )
                    if _tags_display:
                        st.markdown(f'<div style="margin-bottom:10px">{_tags_display}</div>', unsafe_allow_html=True)

                    _tc1, _tc2 = st.columns([3, 1])
                    with _tc1:
                        _tag_choices = [t for t in _TAG_PRESETS if t not in _tag_list]
                        _tag_to_add = st.selectbox("Ajouter une étiquette", [""] + _tag_choices, key=f"tag_add_{_sel_db_id}")
                    with _tc2:
                        st.markdown("<div style='height:26px'></div>", unsafe_allow_html=True)
                        if st.button("Ajouter", key=f"tag_add_btn_{_sel_db_id}", use_container_width=True):
                            if _tag_to_add:
                                _tag_list.append(_tag_to_add)
                                _new_tags = "|".join(_tag_list)
                                _hist_entry = f"{datetime.now().strftime('%d/%m/%Y %H:%M')} — Étiquette ajoutée : {_tag_to_add}"
                                _new_hist = (_current_history + "\n" + _hist_entry).strip()
                                _db_path = str(ROOT / config.db_path)
                                _tc = sqlite3.connect(_db_path)
                                _tc.execute("UPDATE leads SET tags=?, lead_history=?, updated_at=? WHERE id=?",
                                            (_new_tags, _new_hist, datetime.now().isoformat(), int(_lead_id)))
                                _tc.commit(); _tc.close()
                                st.rerun()

                    if _tag_list:
                        _tag_to_rm = st.selectbox("Retirer une étiquette", [""] + _tag_list, key=f"tag_rm_{_sel_db_id}")
                        if _tag_to_rm and st.button("Retirer", key=f"tag_rm_btn_{_sel_db_id}"):
                            _tag_list.remove(_tag_to_rm)
                            _new_tags = "|".join(_tag_list)
                            _db_path = str(ROOT / config.db_path)
                            _tc = sqlite3.connect(_db_path)
                            _tc.execute("UPDATE leads SET tags=?, updated_at=? WHERE id=?",
                                        (_new_tags, datetime.now().isoformat(), int(_lead_id)))
                            _tc.commit(); _tc.close()
                            st.rerun()

                    # ── Notes ───────────────────────────────────────────
                    _new_notes = st.text_area(
                        "Notes de suivi",
                        value=_current_notes,
                        placeholder="Ex: 1er appel le 20/04, rappel prévu le 25/04...",
                        height=100,
                        key=f"fiche_notes_{_sel_db_id}",
                    )

                    if st.button("Enregistrer le suivi", key=f"fiche_save_suivi_{_sel_db_id}", type="primary", use_container_width=True):
                        if _lead_id:
                            _save_status = st.session_state.get(f"fiche_call_st_{_sel_db_id}", _new_status)
                            _save_notes = st.session_state.get(f"fiche_notes_{_sel_db_id}", _new_notes)
                            _db_path = str(ROOT / config.db_path)
                            _suivi_conn = sqlite3.connect(_db_path)
                            _suivi_conn.execute(
                                "UPDATE leads SET call_status=?, lead_notes=?, updated_at=? WHERE id=?",
                                (_save_status, _save_notes, datetime.now().isoformat(), int(_lead_id)),
                            )
                            _suivi_conn.commit()
                            _suivi_conn.close()
                            st.success("Suivi enregistré.")
                            st.rerun()

                    # ── Historique ──────────────────────────────────────
                    st.markdown(
                        '<div style="font-size:13px;font-weight:700;color:#E87B2A;margin:20px 0 8px;letter-spacing:1px">HISTORIQUE</div>',
                        unsafe_allow_html=True,
                    )

                    if _current_history:
                        _hist_html = ""
                        for _h_line in _current_history.strip().split("\n"):
                            if _h_line.strip():
                                _h_parts = _h_line.split(" — ", 1)
                                _h_date = _h_parts[0] if len(_h_parts) > 1 else ""
                                _h_text = _h_parts[1] if len(_h_parts) > 1 else _h_line
                                _hist_html += (
                                    f'<div style="display:flex;gap:10px;padding:6px 12px;background:#0F1016;border-radius:6px;margin-bottom:3px;border-left:2px solid #E87B2A44">'
                                    f'<span style="font-size:10px;color:#4A4D58;white-space:nowrap;min-width:100px">{_h_date}</span>'
                                    f'<span style="font-size:11px;color:#C8C2BB">{_h_text}</span></div>'
                                )
                        st.markdown(_hist_html, unsafe_allow_html=True)
                    else:
                        st.markdown('<div style="color:#4A4D58;font-size:12px;padding:8px">Aucun historique</div>', unsafe_allow_html=True)

                    # ── Réponses email ─────────────────────────────────
                    if _email_d and is_gmail_configured():
                        if st.button("Vérifier les réponses email", key=f"check_replies_{_sel_db_id}", use_container_width=True):
                            with st.spinner("Vérification IMAP..."):
                                _replies = check_replies([_email_d])
                            if _replies:
                                _hist_additions = []
                                for _rp in _replies:
                                    _rp_entry = f"{_rp['date']} — Réponse reçue : {_rp['subject']}"
                                    if _rp_entry not in _current_history:
                                        _hist_additions.append(_rp_entry)
                                if _hist_additions:
                                    _new_hist = (_current_history + "\n" + "\n".join(_hist_additions)).strip()
                                    _new_tag_list = list(_tag_list)
                                    if "Réponse reçue" not in _new_tag_list:
                                        _new_tag_list.append("Réponse reçue")
                                    _db_path = str(ROOT / config.db_path)
                                    _rc = sqlite3.connect(_db_path)
                                    _rc.execute("UPDATE leads SET lead_history=?, tags=?, updated_at=? WHERE id=?",
                                                (_new_hist, "|".join(_new_tag_list), datetime.now().isoformat(), int(_lead_id)))
                                    _rc.commit(); _rc.close()
                                    st.success(f"{len(_hist_additions)} nouvelle(s) réponse(s) détectée(s).")
                                    st.rerun()
                                else:
                                    st.info("Aucune nouvelle réponse.")
                            else:
                                st.info("Aucune réponse trouvée.")


                else:
                    st.markdown(
                        '<div style="display:flex;align-items:center;justify-content:center;height:400px;'
                        'color:#4A4D58;font-size:14px;text-align:center">'
                        '<div>← Sélectionne un lead pour afficher sa fiche détaillée</div></div>',
                        unsafe_allow_html=True,
                    )

      except Exception as e:
        log.error("Erreur onglet Fiches Leads", exc_info=True)
        st.error(f"Erreur : {e}")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 — Configuration / Setup
# ══════════════════════════════════════════════════════════════════════════════
with tab6:
    st.title("Configuration")

    def _save_env_keys(**kwargs):
        env_path = ROOT / ".env"
        lines = []
        if env_path.exists():
            for line in env_path.read_text(encoding="utf-8").splitlines():
                s = line.strip()
                if not s or s.startswith("#"):
                    lines.append(line)
                    continue
                k = s.split("=", 1)[0]
                if k not in kwargs:
                    lines.append(line)
        for k, v in kwargs.items():
            if v.strip():
                lines.append(f"{k}={v.strip()}")
        env_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        load_dotenv(env_path, override=True)
        import os as _osx
        config.serpapi_key   = _osx.getenv("SERPAPI_KEY",        "")
        config.pagespeed_key = _osx.getenv("PAGESPEED_API_KEY",  "")
        config.anthropic_key = _osx.getenv("ANTHROPIC_API_KEY",  "")

    # ── Bannière statut global ────────────────────────────────────────────────
    _cfg_all = load_user_config()
    _serp_ok   = bool(config.serpapi_key)
    _gmail_ok  = bool(_cfg_all.get("gmail_address", "").strip() and _cfg_all.get("gmail_app_password", "").strip())
    _phone_ok  = bool(_cfg_all.get("user_phone", "").strip())
    _voip_ok   = is_twilio_configured() or is_telnyx_configured()
    _crm_ok    = bool([k for k in PUSH_CAPABLE_CRMS if is_connected(k)])

    _status_items = [
        ("API Scraping", _serp_ok),
        ("Email", _gmail_ok),
        ("Telephone", _phone_ok),
        ("VoIP", _voip_ok),
        ("CRM", _crm_ok),
    ]
    _status_html = ''.join(
        f'<span style="display:inline-block;background:{"#0D1A0F" if ok else "#1A1200"};border:1px solid {"#1A3320" if ok else "#3A2E00"};'
        f'border-radius:6px;padding:4px 12px;margin:0 4px 4px 0;font-size:11px;color:{"#2ECC71" if ok else "#E8C32A"};font-weight:600">'
        f'{"✓" if ok else "○"} {lbl}</span>'
        for lbl, ok in _status_items
    )
    st.markdown(f'<div style="margin-bottom:20px">{_status_html}</div>', unsafe_allow_html=True)

    # ── Sous-onglets par thème ────────────────────────────────────────────────
    _cfg_tab_api, _cfg_tab_phone, _cfg_tab_email, _cfg_tab_crm, _cfg_tab_system = st.tabs([
        "API & Scraping",
        "Telephone",
        "Email",
        "CRM",
        "Systeme",
    ])

    # ╔══════════════════════════════════════════════════════════════════════╗
    # ║  API & SCRAPING                                                     ║
    # ╚══════════════════════════════════════════════════════════════════════╝
    with _cfg_tab_api:
        st.markdown('<div class="section-lbl">Cles API</div>', unsafe_allow_html=True)

        c1, c2 = st.columns(2)
        with c1:
            inp_serpapi = st.text_input(
                "SERPAPI_KEY — Google Maps ✱ obligatoire",
                value=config.serpapi_key,
                type="password",
                key="cfg_serpapi",
                help="Clé SerpAPI pour le scraping Google Maps. Obtenir sur serpapi.com."
            )
        with c2:
            inp_pagespeed = st.text_input(
                "PAGESPEED_API_KEY — Google PageSpeed (optionnel)",
                value=config.pagespeed_key,
                type="password",
                key="cfg_pagespeed",
                help="Pour les scores de vitesse mobile/desktop. Obtenir sur console.cloud.google.com."
            )
        inp_anthropic = st.text_input(
            "ANTHROPIC_API_KEY (optionnel)",
            value=config.anthropic_key,
            type="password",
            key="cfg_anthropic",
        )

        if st.button("Enregistrer les cles", type="primary", key="btn_cfg_save"):
            _save_env_keys(
                SERPAPI_KEY=inp_serpapi,
                PAGESPEED_API_KEY=inp_pagespeed,
                ANTHROPIC_API_KEY=inp_anthropic,
            )
            st.success("Clés enregistrées et actives.")
            st.rerun()

        st.divider()

        st.markdown('<div class="section-lbl">Sources de donnees</div>', unsafe_allow_html=True)
        st.markdown(
            """
            | Source | Type | Licence | Données |
            |--------|------|---------|---------|
            | **Google Maps** (SerpAPI) | API payante | Usage commercial OK | Nom, adresse, téléphone, site web, avis |
            | **Registre National** (recherche-entreprises.api.gouv.fr) | API gratuite | Licence ouverte Etalab | Nom, adresse, SIREN, NAF, dirigeants, effectifs |
            | **Dirigeants** (recherche-entreprises.api.gouv.fr) | API gratuite | Licence ouverte Etalab | Dirigeant, forme juridique, date création |
            | **Géocodage** (geo.api.gouv.fr) | API gratuite | Licence ouverte Etalab | Communes voisines, coordonnées GPS |
            | **Google PageSpeed** | API gratuite | Google ToS | Scores mobile/desktop |

            Toutes les sources utilisées sont **légales et autorisent l'usage commercial**.
            """,
            unsafe_allow_html=True,
        )

    # ╔══════════════════════════════════════════════════════════════════════╗
    # ║  TELEPHONE                                                          ║
    # ╚══════════════════════════════════════════════════════════════════════╝
    with _cfg_tab_phone:
        # Numéro personnel
        st.markdown('<div class="section-lbl">Ton numero</div>', unsafe_allow_html=True)
        st.caption("Numéro sur lequel tu recevras les appels VoIP (Telnyx/Twilio t'appellent d'abord, puis te connectent au lead).")

        _cfg_phone = load_user_config()
        _user_phone = _cfg_phone.get("user_phone", "")

        if _user_phone:
            st.markdown(
                f'<div style="background:#0D1A0F;border:1px solid #1A3320;border-radius:8px;padding:10px 16px;margin-bottom:12px">'
                f'<span style="color:#2ECC71;font-size:12px;font-weight:600">✓ {_user_phone}</span></div>',
                unsafe_allow_html=True,
            )

        with st.expander("Configurer le numero" if not _user_phone else "Modifier le numero"):
            _new_phone = st.text_input("Ton numéro de téléphone", value=_user_phone, key="cfg_user_phone", placeholder="+33 6 12 34 56 78")
            if st.button("Enregistrer", key="cfg_save_phone", type="primary"):
                save_user_config({"user_phone": _new_phone.strip()})
                st.success("Numéro enregistré.")
                st.rerun()

        st.divider()

        # Telnyx
        st.markdown('<div class="section-lbl">Telnyx — VoIP (~0.006eur/min)</div>', unsafe_allow_html=True)
        st.caption("Moins cher, streaming audio natif. Nécessite un email professionnel pour l'inscription.")

        _cfg_tel = load_user_config()
        _tel_api = _cfg_tel.get("telnyx_api_key", "")
        _tel_conn = _cfg_tel.get("telnyx_connection_id", "")
        _tel_num = _cfg_tel.get("telnyx_phone_number", "")

        if _tel_api and _tel_conn and _tel_num:
            st.markdown(
                '<div style="background:#0D1A0F;border:1px solid #1A3320;border-radius:8px;padding:10px 16px;margin-bottom:12px">'
                '<span style="color:#2ECC71;font-size:12px;font-weight:600">✓ Telnyx configuré</span></div>',
                unsafe_allow_html=True,
            )

        with st.expander("Configurer Telnyx" if not _tel_api else "Modifier Telnyx"):
            st.markdown(
                '<div style="font-size:11px;color:#6B7280;margin-bottom:12px">'
                '1. Crée un compte sur <b>telnyx.com</b><br>'
                '2. Récupère ta <b>clé API</b> dans API Keys<br>'
                '3. Crée une <b>SIP Connection</b> (Call Control) et copie son ID<br>'
                '4. Achète un <b>numéro français</b> (+33) dans Numbers</div>',
                unsafe_allow_html=True,
            )
            _new_tel_api = st.text_input("Clé API Telnyx", value=_tel_api, key="cfg_telnyx_api", type="password")
            _new_tel_conn = st.text_input("Connection ID", value=_tel_conn, key="cfg_telnyx_conn", placeholder="ex: 1494404757889422000")
            _new_tel_num = st.text_input("Numéro Telnyx (+33...)", value=_tel_num, key="cfg_telnyx_num", placeholder="+33 1 23 45 67 89")
            if st.button("Enregistrer Telnyx", key="cfg_save_telnyx", type="primary"):
                save_user_config({
                    "telnyx_api_key": _new_tel_api.strip(),
                    "telnyx_connection_id": _new_tel_conn.strip(),
                    "telnyx_phone_number": _new_tel_num.strip(),
                })
                st.success("Configuration Telnyx enregistrée.")
                st.rerun()

        st.divider()

        # Twilio
        st.markdown('<div class="section-lbl">Twilio — VoIP (~0.013eur/min)</div>', unsafe_allow_html=True)
        st.caption("Inscription avec email perso (Gmail OK). Documentation très complète.")

        _cfg_tw = load_user_config()
        _tw_sid = _cfg_tw.get("twilio_account_sid", "")
        _tw_tok = _cfg_tw.get("twilio_auth_token", "")
        _tw_num = _cfg_tw.get("twilio_phone_number", "")

        if _tw_sid and _tw_tok and _tw_num:
            st.markdown(
                '<div style="background:#0D1A0F;border:1px solid #1A3320;border-radius:8px;padding:10px 16px;margin-bottom:12px">'
                '<span style="color:#2ECC71;font-size:12px;font-weight:600">✓ Twilio configuré</span></div>',
                unsafe_allow_html=True,
            )

        with st.expander("Configurer Twilio" if not _tw_sid else "Modifier Twilio"):
            st.markdown(
                '<div style="font-size:11px;color:#6B7280;margin-bottom:12px">'
                '1. Crée un compte sur <b>twilio.com</b> (email perso OK)<br>'
                '2. Récupère ton <b>Account SID</b> et <b>Auth Token</b> dans la console<br>'
                '3. Achète un <b>numéro français</b> (+33) dans Phone Numbers</div>',
                unsafe_allow_html=True,
            )
            _new_tw_sid = st.text_input("Account SID", value=_tw_sid, key="cfg_twilio_sid", placeholder="ACxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx")
            _new_tw_tok = st.text_input("Auth Token", value=_tw_tok, key="cfg_twilio_tok", type="password")
            _new_tw_num = st.text_input("Numéro Twilio (+33...)", value=_tw_num, key="cfg_twilio_num", placeholder="+33 1 23 45 67 89")
            if st.button("Enregistrer Twilio", key="cfg_save_twilio", type="primary"):
                save_user_config({
                    "twilio_account_sid": _new_tw_sid.strip(),
                    "twilio_auth_token": _new_tw_tok.strip(),
                    "twilio_phone_number": _new_tw_num.strip(),
                })
                st.success("Configuration Twilio enregistrée.")
                st.rerun()

    # ╔══════════════════════════════════════════════════════════════════════╗
    # ║  EMAIL                                                              ║
    # ╚══════════════════════════════════════════════════════════════════════╝
    with _cfg_tab_email:
        st.markdown('<div class="section-lbl">Gmail SMTP</div>', unsafe_allow_html=True)
        st.caption("Envoie des emails directement depuis les fiches leads via ton compte Gmail.")

        _cfg_mail = load_user_config()
        _gmail_addr = _cfg_mail.get("gmail_address", "")
        _gmail_pass = _cfg_mail.get("gmail_app_password", "")

        if _gmail_addr and _gmail_pass:
            st.markdown(
                f'<div style="background:#0D1A0F;border:1px solid #1A3320;border-radius:8px;padding:10px 16px;margin-bottom:12px">'
                f'<span style="color:#2ECC71;font-size:12px;font-weight:600">✓ Gmail configuré — {_gmail_addr}</span></div>',
                unsafe_allow_html=True,
            )

        with st.expander("Configurer Gmail" if not (_gmail_addr and _gmail_pass) else "Modifier la configuration"):
            st.markdown(
                "**Comment obtenir un mot de passe d'application :**\n\n"
                "1. Va sur [myaccount.google.com/apppasswords](https://myaccount.google.com/apppasswords)\n"
                "2. Connecte-toi à ton compte Google\n"
                "3. Donne un nom (ex: LeadsEngine) et clique **Créer**\n"
                "4. Copie le mot de passe de 16 caractères généré\n"
                "5. Colle-le ci-dessous\n\n"
                "⚠️ La validation en 2 étapes doit être activée sur ton compte Google."
            )
            _mc1, _mc2 = st.columns(2)
            with _mc1:
                _new_gmail = st.text_input("Adresse Gmail", value=_gmail_addr, key="cfg_gmail_addr", placeholder="ton.email@gmail.com")
            with _mc2:
                _new_gmail_pass = st.text_input("Mot de passe d'application", value=_gmail_pass, key="cfg_gmail_pass", type="password", placeholder="xxxx xxxx xxxx xxxx")
            if st.button("Enregistrer Gmail", key="cfg_save_gmail", type="primary"):
                save_user_config({"gmail_address": _new_gmail, "gmail_app_password": _new_gmail_pass})
                st.success("Configuration Gmail enregistrée.")
                st.rerun()

    # ╔══════════════════════════════════════════════════════════════════════╗
    # ║  CRM                                                                ║
    # ╚══════════════════════════════════════════════════════════════════════╝
    with _cfg_tab_crm:
        st.markdown('<div class="section-lbl">CRM par defaut</div>', unsafe_allow_html=True)

        _current_crm = load_user_config().get("crm", "default")
        _connected_crms = [k for k in PUSH_CAPABLE_CRMS if is_connected(k)]
        _crm_keys = ["default"] + list(CRM_MAPPINGS.keys())
        _crm_labels = {"default": "Aucun (export Excel/CSV standard)"}
        _crm_labels.update({k: v["label"] + (" — Connecté" if k in _connected_crms else "") for k, v in CRM_MAPPINGS.items()})

        _crm_index = _crm_keys.index(_current_crm) if _current_crm in _crm_keys else 0
        _new_crm = st.selectbox(
            "CRM principal",
            options=_crm_keys,
            index=_crm_index,
            format_func=lambda x: _crm_labels.get(x, x),
            key="cfg_crm_select",
            help="Le CRM sélectionné sera proposé en priorité dans l'export des recherches.",
        )
        if _new_crm != _current_crm:
            set_crm(_new_crm)
            st.rerun()

        if _connected_crms:
            _connected_labels = ", ".join(PUSH_CAPABLE_CRMS[k]["label"] for k in _connected_crms)
            st.markdown(
                f'<div style="background:#0A1A10;border:1px solid #1A4228;border-left:3px solid #2ECC71;'
                f'border-radius:6px;padding:10px 14px;margin:12px 0;font-size:12px;color:#7DFAB8">'
                f'CRM connectés : <b>{_connected_labels}</b> — push API disponible</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div style="background:#1A1200;border:1px solid #3A2E00;border-left:3px solid #E8C32A;'
                'border-radius:6px;padding:10px 14px;margin:12px 0;font-size:12px;color:#F5D87A">'
                'Aucun CRM connecté — connecte un CRM ci-dessous pour activer le push direct.</div>',
                unsafe_allow_html=True,
            )

        st.divider()

        st.markdown('<div class="section-lbl">Connexions</div>', unsafe_allow_html=True)

        for _crm_key, _crm_info in PUSH_CAPABLE_CRMS.items():
            _is_conn = is_connected(_crm_key)
            _is_default = (_current_crm == _crm_key)

            with st.expander(f"{_crm_info['label']} {':white_check_mark:' if _is_conn else ''}"):
                if not _is_conn:
                    st.markdown(f"**Comment connecter {_crm_info['label']} :**\n\n{_crm_info['guide']}")

                _cfg = load_user_config()
                _main_key = _crm_info["config_key"]

                if _crm_info["auth_type"] == "oauth":
                    if not _is_conn:
                        if _crm_key == "hubspot":
                            _auth_url = hubspot_auth_url()
                            _placeholder = "eu1-xxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx"
                        elif _crm_key == "salesforce":
                            _auth_url = salesforce_auth_url()
                            _placeholder = "aPrxJq...long_code..."
                        else:
                            _auth_url = ""
                            _placeholder = ""
                        st.link_button(f"Connecter {_crm_info['label']}", _auth_url, type="primary", use_container_width=True)
                        st.caption("Après autorisation, copie le code depuis la barre d'adresse (après `code=`) :")
                        _oauth_code = st.text_input(
                            "Code d'autorisation",
                            placeholder=_placeholder,
                            key=f"push_oauth_code_{_crm_key}",
                        )
                        if _oauth_code:
                            with st.spinner("Connexion en cours…"):
                                if _crm_key == "hubspot":
                                    _oauth_ok, _oauth_msg = hubspot_exchange_code(_oauth_code)
                                elif _crm_key == "salesforce":
                                    _oauth_ok, _oauth_msg = salesforce_exchange_code(_oauth_code)
                                else:
                                    _oauth_ok, _oauth_msg = False, "CRM non supporté"
                            if _oauth_ok:
                                st.success(_oauth_msg)
                                st.rerun()
                            else:
                                st.error(_oauth_msg)
                else:
                    if not _is_conn:
                        _token_val = st.text_input(
                            _crm_info["auth_label"],
                            value=_cfg.get(_main_key, ""),
                            type="password",
                            key=f"push_{_crm_key}_key",
                        )
                        if st.button("Enregistrer", key=f"push_save_{_crm_key}", type="primary"):
                            save_user_config({_main_key: _token_val})
                            st.success(f"{_crm_info['label']} connecté.")
                            st.rerun()

                if _is_conn:
                    _tc1, _tc2 = st.columns(2)
                    with _tc1:
                        if st.button("Tester la connexion", key=f"push_test_{_crm_key}", use_container_width=True):
                            with st.spinner("Test…"):
                                _ok, _msg = test_connection(_crm_key)
                            if _ok:
                                st.success(_msg)
                            else:
                                st.error(_msg)
                    with _tc2:
                        if st.button("Déconnecter", key=f"push_disconnect_{_crm_key}", type="secondary", use_container_width=True):
                            _keys_to_clear = {_main_key: ""}
                            for _ef in _crm_info.get("extra_fields", []):
                                _keys_to_clear[_ef] = ""
                            if _crm_key == "hubspot":
                                _keys_to_clear["hubspot_refresh_token"] = ""
                            elif _crm_key == "salesforce":
                                _keys_to_clear["sf_refresh_token"] = ""
                                _keys_to_clear["sf_instance_url"] = ""
                            save_user_config(_keys_to_clear)
                            st.success(f"{_crm_info['label']} déconnecté.")
                            st.rerun()

    # ╔══════════════════════════════════════════════════════════════════════╗
    # ║  SYSTEME                                                            ║
    # ╚══════════════════════════════════════════════════════════════════════╝
    with _cfg_tab_system:
        # Licence
        st.markdown('<div class="section-lbl">Licence</div>', unsafe_allow_html=True)
        _lic_tier = get_tier()
        _lic_key = get_license_key()
        _lic_color = "#E87B2A" if _lic_tier == "pro" else "#4A4D58"
        st.markdown(
            f'<div style="font-family:IBM Plex Mono,monospace;font-size:12px;color:#C8C2BB;margin-bottom:8px">'
            f'Licence : <span style="background:{_lic_color};color:#0D0E11;font-size:10px;font-weight:800;'
            f'padding:2px 8px;border-radius:3px">{_lic_tier.upper()}</span>'
            f'</div>'
            f'<div style="font-family:IBM Plex Mono,monospace;font-size:11px;color:#4A4D58;margin-bottom:12px">'
            f'Clé : {_lic_key}</div>',
            unsafe_allow_html=True,
        )
        _c_lic1, _c_lic2 = st.columns([2, 1])
        with _c_lic1:
            _new_key = st.text_input("Changer de clé", key="cfg_license_key", placeholder="LE-XXX-XXXX-XXXX")
        with _c_lic2:
            st.markdown("<div style='height:26px'></div>", unsafe_allow_html=True)
            if st.button("Activer", key="btn_cfg_activate", type="primary"):
                if _new_key.strip():
                    _ok_lic, _msg_lic = activate_license(_new_key)
                    if _ok_lic:
                        st.success(_msg_lic)
                        time.sleep(1)
                        st.rerun()
                    else:
                        st.error(_msg_lic)

        st.divider()

        # Fichiers & dossiers
        st.markdown('<div class="section-lbl">Fichiers de donnees</div>', unsafe_allow_html=True)
        st.markdown(f'<div style="font-family:IBM Plex Mono,monospace;font-size:11px;color:#4A4D58;margin-bottom:12px">Dossier racine : {ROOT}</div>', unsafe_allow_html=True)

        _items = [
            ("leads.db",       ROOT / config.db_path, False),
            ("crm/",           ROOT / "crm",           False),
            ("analyses_a2/",   ROOT / "analyses_a2",   False),
            (".env",           ROOT / ".env",           False),
        ]
        for _lbl, _path, _warn_only in _items:
            _ok    = _path.exists()
            _color = "#2ECC71" if _ok else ("#E8C32A" if _warn_only else "#E84C4C")
            _icon  = "✓" if _ok else "⚠" if _warn_only else "✗"
            st.markdown(
                f'<div style="font-family:IBM Plex Mono,monospace;font-size:12px;margin-bottom:6px">'
                f'<span style="color:{_color};font-weight:700">{_icon}</span>&nbsp;&nbsp;'
                f'<span style="color:#C8C2BB">{_lbl}</span>&nbsp;&nbsp;'
                f'<span style="color:#4A4D58">{_path}</span></div>',
                unsafe_allow_html=True,
            )

        if st.button("Créer les dossiers manquants", key="btn_cfg_mkdir"):
            (ROOT / "crm").mkdir(parents=True, exist_ok=True)
            (ROOT / "analyses_a2").mkdir(parents=True, exist_ok=True)
            LeadQueue(str(ROOT / config.db_path))
            st.success("Dossiers et base de données initialisés.")
            st.rerun()

        st.divider()

        # Mise à jour
        st.markdown('<div class="section-lbl">Mise a jour</div>', unsafe_allow_html=True)

        _local_v = get_local_version(ROOT)
        st.markdown(f'<div style="font-family:IBM Plex Mono,monospace;font-size:12px;color:#C8C2BB;margin-bottom:12px">Version actuelle : <b>v{_local_v}</b></div>', unsafe_allow_html=True)

        if st.button("Vérifier les mises à jour", key="btn_check_update"):
            with st.spinner("Vérification en cours…"):
                update = check_update(ROOT)
            if update is None:
                st.success("L'application est à jour.")
            else:
                st.session_state["_pending_update"] = update

        _upd = st.session_state.get("_pending_update")
        if _upd:
            _size_mb = _upd.get("size", 0) / (1024 * 1024)
            st.warning(f"Nouvelle version disponible : **v{_upd['version']}**  ({_size_mb:.0f} Mo)")
            if _upd.get("changelog"):
                with st.expander("Notes de version"):
                    st.markdown(_upd["changelog"])

            if st.button("Installer la mise à jour", type="primary", key="btn_install_update"):
                _progress = st.progress(0, text="Téléchargement…")
                _ok = download_and_install(
                    ROOT, _upd,
                    progress_callback=lambda p: _progress.progress(p, text=f"Téléchargement… {p*100:.0f}%"),
                )
                if _ok:
                    _progress.progress(1.0, text="Téléchargement terminé !")
                    st.success("Mise à jour téléchargée — fermeture et installation…")
                    st.session_state.pop("_pending_update", None)
                    time.sleep(1)
                    launch_update_and_quit(ROOT)
                else:
                    st.error("Échec du téléchargement. Vérifie ta connexion et réessaie.")
